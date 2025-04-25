# -*- coding: utf-8 -*-
"""
Processes an Excel workbook (.xlsx) to extract routing entities (VQs, Skills,
VAGs, Skill Expressions), compares them against data fetched from Genesys
configuration APIs, and reports differences in new sheets within the workbook.

Handles strikethrough formatting: if an item appears both with and without
strikethrough in the source sheet, it's treated as effectively 'present'
(not struck out) for comparison purposes.

The 'Skill Expr' output sheet includes separate 'Expression' and
'Ideal Expression' columns for reference, while comparison uses the
'Concatenated Key'. The 'Skill_exprs Comparison' sheet also includes
these separate columns.

Calculates the maximum numeric ID found SEPARATELY for DN (VQ) data and
Agent Group (Skill/VAG/Expr) data and stores them in a 'Metadata' sheet.

Configuration (file paths, API URLs, sheet layout details) is loaded from
a 'config.ini' file expected in the same directory.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import cell as openpyxl_cell_utils
import re
import requests
import logging
import shutil
import os
import configparser
from typing import Dict, Any, Optional, Tuple, Set, List

# --- Constants ---
CONFIG_FILE = 'config.ini'
LOG_FILE = 'log.txt'
METADATA_SHEET_NAME = "Metadata" # Name for the sheet storing max ID
MAX_DN_ID_LABEL_CELL = "A1"      # Cell Coordinate for Max DN ID Label
MAX_DN_ID_VALUE_CELL = "B1"      # Cell Coordinate for Max DN ID Value
MAX_AG_ID_LABEL_CELL = "A2"      # Cell Coordinate for Max Agent Group ID Label
MAX_AG_ID_VALUE_CELL = "B2"      # Cell Coordinate for Max Agent Group ID Value


# --- Logging Setup ---
# Configure logging to file and console
logging.basicConfig(
    level=logging.INFO, # Set root logger level (can be DEBUG for more detail)
    format='%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, mode='w'), # Overwrite log file each run
        logging.StreamHandler() # Log to console (stderr by default)
    ]
)
# Adjust console handler level if needed (e.g., only show INFO+)
for handler in logging.getLogger().handlers:
    if isinstance(handler, logging.StreamHandler):
        handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(levelname)s: %(message)s') # Simpler format for console
        handler.setFormatter(formatter)


# --- Configuration Loading ---
def load_configuration(config_path: str) -> Dict[str, Any]:
    """
    Loads settings from the specified INI configuration file.

    Args:
        config_path: Path to the configuration file (e.g., 'config.ini').

    Returns:
        A dictionary containing the loaded configuration settings.

    Raises:
        FileNotFoundError: If the configuration file does not exist.
        ValueError: If the configuration file has missing sections/options,
                    invalid values, or other parsing errors.
    """
    logging.info(f"Loading configuration from: {config_path}")
    if not os.path.exists(config_path):
        logging.error(f"Configuration file '{config_path}' not found.")
        raise FileNotFoundError(f"Configuration file '{config_path}' not found.")

    config = configparser.ConfigParser(interpolation=None) # Disable interpolation for literal values
    try:
        config.read(config_path, encoding='utf-8') # Specify encoding

        # Define required and optional settings
        settings = {}
        required_sections = {
            'Files': ['source_file'],
            'API': ['dn_url', 'agent_group_url'],
            'SheetLayout': ['ideal_agent_header_text', 'ideal_agent_fallback_cell', 'vag_extraction_sheet']
        }
        # Define optional settings as tuples: (section, key, fallback_value)
        optional_settings_spec = {('API', 'timeout', 15)} # Note: key is 'timeout' here

        # Read required settings
        for section, keys in required_sections.items():
            if not config.has_section(section):
                 raise ValueError(f"Missing section '{section}' in configuration file.")
            for key in keys:
                 if not config.has_option(section, key):
                      raise ValueError(f"Missing option '{key}' in section '[{section}]'.")
                 # Read value as string initially
                 settings[key] = config.get(section, key)

        # Read optional settings
        for section, key, fallback in optional_settings_spec:
            if config.has_option(section, key):
                # Perform type conversion if needed
                if key == 'timeout':
                    # Store with the key 'api_timeout' for clarity later
                    settings['api_timeout'] = config.getint(section, key)
                else:
                    settings[key] = config.get(section, key)
            else:
                # Use fallback value if option is missing, store with 'api_timeout' key
                if key == 'timeout':
                    settings['api_timeout'] = fallback
                else:
                    settings[key] = fallback
                logging.debug(f"Optional setting '{key}' not found in '[{section}]', using fallback: {fallback}")


        logging.info("Configuration loaded successfully.")

        # --- Basic Validation ---
        # Check if source file is .xlsx
        if not settings['source_file'].lower().endswith('.xlsx'):
             msg = f"Source file '{settings['source_file']}' in config must be an .xlsx file."
             logging.error(msg)
             raise ValueError(msg)

        # Validate fallback cell format (e.g., "C2")
        try:
            openpyxl_cell_utils.coordinate_to_tuple(settings['ideal_agent_fallback_cell'])
        except openpyxl_cell_utils.IllegalCharacterError:
            msg = f"Invalid format for 'ideal_agent_fallback_cell' in config: {settings['ideal_agent_fallback_cell']}"
            logging.error(msg)
            raise ValueError(msg)

        return settings

    except (configparser.NoSectionError, configparser.NoOptionError, ValueError) as e:
        # Catch specific config errors raised above or by configparser
        logging.error(f"Configuration error in '{config_path}': {e}")
        raise ValueError(f"Configuration error in '{config_path}': {e}")
    except Exception as e:
        # Catch any other unexpected errors during reading/parsing
        logging.error(f"Unexpected error reading configuration '{config_path}': {e}", exc_info=True)
        raise ValueError(f"Unexpected error reading configuration '{config_path}': {e}")


# --- Excel Utilities ---
def copy_cell_style(source_cell: openpyxl.cell.Cell, target_cell: openpyxl.cell.Cell):
    """
    Copies font, fill, alignment, and number format style from source_cell to target_cell.

    Args:
        source_cell: The cell to copy style from.
        target_cell: The cell to copy style to.
    """
    if source_cell.has_style:
        # Copy Font properties
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        # Copy Fill properties
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        # Copy Alignment properties
        if source_cell.alignment:
            target_cell.alignment = openpyxl.styles.Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
        # Copy Number format
        target_cell.number_format = source_cell.number_format
        # Copy Border (more complex, omitted for brevity but could be added)
        # target_cell.border = Border(...)
    else:
        # Apply default styles if source has no specific style applied
        target_cell.font = Font()
        target_cell.fill = PatternFill()
        target_cell.alignment = openpyxl.styles.Alignment()
        target_cell.number_format = 'General'


def identify_ideal_agent_column(sheet: openpyxl.worksheet.worksheet.Worksheet, config: Dict[str, Any]) -> Optional[int]:
    """
    Identifies the column index for the 'Ideal Agent' column based on config.
    Searches header row 1 (specifically Columns C & D) first, then checks a fallback cell.

    Args:
        sheet: The openpyxl worksheet object.
        config: The loaded configuration dictionary.

    Returns:
        The column index (1-based) if found, otherwise None.
    """
    header_text = config['ideal_agent_header_text']
    fallback_cell_coord = config['ideal_agent_fallback_cell']
    logging.debug(f"Identifying '{header_text}' column in sheet: {sheet.title}")

    # 1. Check headers in row 1, specifically columns C and D
    for col_idx in [3, 4]:  # Column C=3, Column D=4
        # Check if sheet has enough columns before accessing cell
        if col_idx <= sheet.max_column:
            cell_value = sheet.cell(row=1, column=col_idx).value
            # Check if cell has a value and contains the header text
            if cell_value and header_text in str(cell_value):
                logging.debug(f"Found '{header_text}' in header row 1 at column {col_idx}")
                return col_idx
        else:
            logging.debug(f"Sheet '{sheet.title}' has only {sheet.max_column} columns, cannot check column {col_idx} in header.")

    # 2. Check specific fallback cell from config if not found in header cols C/D
    try:
        # Convert cell coordinate like "C2" to row/col index
        col_str, row_str = openpyxl_cell_utils.coordinate_to_tuple(fallback_cell_coord)
        fallback_col_idx = openpyxl_cell_utils.column_index_from_string(col_str)
        fallback_row_idx = int(row_str)

        # Check if fallback cell is within sheet bounds
        if fallback_row_idx <= sheet.max_row and fallback_col_idx <= sheet.max_column:
            cell_value_fallback = sheet.cell(row=fallback_row_idx, column=fallback_col_idx).value
            if cell_value_fallback and header_text in str(cell_value_fallback):
                logging.debug(f"Found '{header_text}' at fallback cell {fallback_cell_coord} (Col {fallback_col_idx})")
                return fallback_col_idx
        else:
            logging.warning(f"Fallback cell '{fallback_cell_coord}' is outside the bounds of sheet '{sheet.title}'.")

    except Exception as e:
         # Catch potential errors during coordinate conversion or cell access
         logging.warning(f"Could not parse or check fallback cell '{fallback_cell_coord}': {e}")

    # 3. If not found by either method
    logging.warning(f"'{header_text}' column not found using header search (Cols C/D) or fallback cell '{fallback_cell_coord}' in sheet: {sheet.title}.")
    return None

def extract_skills(expression: str) -> list[str]:
    """
    Extracts potential skill names (alphanumeric + underscore) from a skill
    expression string. Looks for patterns like 'SkillName>5'.

    Args:
        expression: The skill expression string.

    Returns:
        A list of extracted skill names.
    """
    # Regex finds words (alphanumeric or underscore) followed immediately by '>' and one or more digits
    skills = re.findall(r'\b([a-zA-Z0-9_]+)(?=>\d+)', expression)
    logging.debug(f"Extracted skills {skills} from expression '{expression}'")
    return skills


# --- API Interaction ---
def fetch_api_data(config: Dict[str, Any]) -> Tuple[Dict[str, Dict[str, Any]], int, int]:
    """
    Fetches routing entity data from APIs specified in the configuration.
    Calculates the maximum numeric ID found SEPARATELY for DN and Agent Group sources.
    Stores detailed info for skill expressions.

    Args:
        config: The loaded configuration dictionary.

    Returns:
        A tuple containing:
        - api_data: Dictionary containing the fetched API data. Structure varies by key.
        - max_dn_id: The highest numeric ID found in the DN API data (or 0).
        - max_ag_id: The highest numeric ID found in the Agent Group API data (or 0).
    """
    logging.info("Fetching API data...")
    # Use correct keys from config dict as stored by load_configuration
    dn_url = config['dn_url']
    agent_group_url = config['agent_group_url']
    timeout = config['api_timeout']

    api_data = {"vqs": {}, "skills": {}, "skill_exprs": {}, "vags": {}}
    # Separate Max ID counters
    max_dn_id = 0
    max_ag_id = 0

    # Initialize JSON variables in case requests fail
    dn_json = []
    ag_json = []

    # Fetch DN (VQ) data
    try:
        logging.debug(f"Fetching DN data from {dn_url} with timeout={timeout}s")
        dn_response = requests.get(dn_url, timeout=timeout)
        dn_response.raise_for_status() # Raise HTTPError for bad responses (4xx or 5xx)
        dn_json = dn_response.json()
        logging.info(f"Successfully fetched DN response ({len(dn_json)} items).")
    except requests.exceptions.Timeout:
         logging.error(f"API request timed out for DN URL after {timeout} seconds.")
         print(f"ERROR: API request timed out for DN URL. Check URL and network.")
         # Continue to try fetching AG data, max_dn_id remains 0
    except requests.exceptions.RequestException as e:
        logging.error(f"API fetch failed for DN URL: {e}")
        print(f"ERROR: Failed to fetch data from DN API. Check URL and network. Details in {LOG_FILE}.")
        # Continue to try fetching AG data, max_dn_id remains 0

    # Fetch Agent Group data
    try:
        logging.debug(f"Fetching Agent Group data from {agent_group_url} with timeout={timeout}s")
        ag_response = requests.get(agent_group_url, timeout=timeout)
        ag_response.raise_for_status()
        ag_json = ag_response.json()
        logging.info(f"Successfully fetched Agent Group response ({len(ag_json)} items).")
    except requests.exceptions.Timeout:
         logging.error(f"API request timed out for Agent Group URL after {timeout} seconds.")
         print(f"ERROR: API request timed out for Agent Group URL. Check URL and network.")
         # Return current api_data and calculated max IDs (max_ag_id will be 0)
         return api_data, max_dn_id, max_ag_id
    except requests.exceptions.RequestException as e:
        logging.error(f"API fetch failed for Agent Group URL: {e}")
        print(f"ERROR: Failed to fetch data from Agent Group API. Check URL and network. Details in {LOG_FILE}.")
         # Return current api_data and calculated max IDs (max_ag_id will be 0)
        return api_data, max_dn_id, max_ag_id


    # --- Process DN (VQ) data ---
    vq_count = 0
    for item in dn_json: # Use the fetched dn_json
        data = item.get('data', {})
        vq_name = data.get('name')
        vq_id = data.get('id') # API might return int or string ID
        # Ensure both name and ID are present
        if vq_name and vq_id is not None:
            # Normalize name: remove spaces and non-breaking spaces (\u00A0)
            normalized_vq = vq_name.replace(" ", "").replace('\u00A0', '')
            id_str = str(vq_id) # Store ID as string for consistency
            api_data["vqs"][normalized_vq] = id_str
            # --- Calculate Max DN ID ---
            # Check if the ID is numeric and update max_dn_id if it's larger
            if id_str.isdigit():
                max_dn_id = max(max_dn_id, int(id_str))
            # --- End Calculate Max DN ID ---
            logging.debug(f"Processed VQ: Name='{normalized_vq}', ID='{id_str}'")
            vq_count += 1
        else:
            logging.warning(f"Skipping DN item due to missing name or id: {item}")
    logging.info(f"Processed {vq_count} VQs from API. Max DN ID found: {max_dn_id}")


    # --- Process Agent Group (Skill, Skill Expr, VAG) data ---
    skill_count, expr_count, vag_count = 0, 0, 0
    skipped_ag_count = 0
    for item in ag_json: # Use the fetched ag_json
        data = item.get('data', {})
        ag_id = data.get('id')
        expression = data.get('expression', '') or '' # Ensure string, default to empty
        ideal_expression = data.get('IdealExpression', '') or '' # Ensure string, default to empty

        # Skip if ID is missing
        if ag_id is None:
            logging.warning(f"Skipping AG item due to missing id: {item}")
            skipped_ag_count += 1
            continue
        ag_id_str = str(ag_id) # Store ID as string

        # --- Calculate Max AG ID ---
        # Check if the ID is numeric and update max_ag_id if it's larger
        if ag_id_str.isdigit():
            max_ag_id = max(max_ag_id, int(ag_id_str))
        # --- End Calculate Max AG ID ---

        # Normalize expressions: remove spaces, add spaces around operators for consistency
        norm_expr = expression.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")
        norm_ideal = ideal_expression.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")

        # Create the concatenated key used for comparison, mirroring sheet processing
        combined_expr = norm_expr
        is_skill_expr = ">" in norm_expr
        if is_skill_expr and norm_ideal: # Only combine if it's a skill expression AND has an ideal part
             combined_expr = f"{norm_expr} {norm_ideal}".strip()

        # Categorize and store based on expression content
        if is_skill_expr: # Skill Expression (contains '>')
            # Store dict with details needed for comparison sheet
            api_data["skill_exprs"][combined_expr] = {
                'id': ag_id_str,
                'expr': norm_expr,   # Store the base expression from API
                'ideal': norm_ideal # Store the ideal part from API
            }
            logging.debug(f"Processed Skill Expr: Key='{combined_expr}', ID='{ag_id_str}'")
            expr_count += 1
        elif "VAG_" in norm_expr: # VAG (check for prefix)
            api_data["vags"][norm_expr] = ag_id_str # Store only ID for VAGs
            logging.debug(f"Processed VAG: Name='{norm_expr}', ID='{ag_id_str}'")
            vag_count += 1
        elif norm_expr: # Potentially a Simple Skill (if not VAG or Skill Expr)
             # Check it contains actual characters, not just operators/spaces
             if re.search(r'[a-zA-Z0-9]', norm_expr):
                 api_data["skills"][norm_expr] = ag_id_str # Store only ID for Skills
                 logging.debug(f"Processed Skill: Name='{norm_expr}', ID='{ag_id_str}'")
                 skill_count += 1
             else:
                 # Log if expression becomes empty/invalid after normalization
                 logging.warning(f"Skipping AG item - skill name seems empty or invalid after normalization: {item}")
                 skipped_ag_count += 1
        else:
             # Log if original expression was empty
             logging.warning(f"Skipping AG item with empty expression: {item}")
             skipped_ag_count += 1

    logging.info(f"Processed Agent Groups from API: Skills={skill_count}, SkillExprs={expr_count}, VAGs={vag_count}. Skipped={skipped_ag_count}.")
    logging.info(f"Finished parsing API data. Max Agent Group ID found: {max_ag_id}")
    # Return the structured data and BOTH max IDs
    return api_data, max_dn_id, max_ag_id


# --- Core Processing Logic ---
def process_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    intermediate_data: Dict[str, Dict[str, Dict[str, Any]]],
    config: Dict[str, Any]
):
    """
    Processes a single sheet to extract routing entities and their strike status,
    updating the intermediate_data dictionary according to the strike-preference rule.
    Stores separate expr/ideal for skill_exprs.

    Args:
        sheet: The openpyxl worksheet object to process.
        intermediate_data: The dictionary holding collected data across sheets.
        config: The loaded configuration dictionary.
    """
    # Define sheets to skip (output, comparison, metadata)
    excluded_sheets = {
        "Skill Expr", "VQ", "VAG", "Skills",
        "Vqs Comparison", "Skills Comparison",
        "Skill_exprs Comparison", "Vags Comparison",
        METADATA_SHEET_NAME # Ensure metadata sheet is skipped
    }
    if sheet.title in excluded_sheets:
        logging.debug(f"Skipping sheet: {sheet.title} (excluded name)")
        return

    logging.info(f"Processing sheet: {sheet.title} (Max Row: {sheet.max_row}, Max Col: {sheet.max_column})")
    ideal_agent_col_idx = identify_ideal_agent_column(sheet, config)
    vag_sheet_name = config['vag_extraction_sheet']

    # Helper for VQ, VAG, Skills (non-skill-expressions)
    def update_intermediate_generic(data_dict: Dict, key: str, current_strike: bool, cell_obj: openpyxl.cell.Cell):
        """Updates the dict for generic items, preferring non-struck entries."""
        if not key: # Skip empty keys
             return
        # Check if item exists
        if key not in data_dict:
            # Add new item
            data_dict[key] = {"strike": current_strike, "style_cell": cell_obj}
            logging.debug(f"Sheet '{sheet.title}' Row {cell_obj.row}: Added new item '{key}' with strike={current_strike}")
        # If item exists, update only if changing strike from True to False
        elif data_dict[key]["strike"] and not current_strike:
            data_dict[key]["strike"] = False
            data_dict[key]["style_cell"] = cell_obj # Use style from non-struck cell
            logging.debug(f"Sheet '{sheet.title}' Row {cell_obj.row}: Updated item '{key}' to strike=False")
        # else: if existing is False, or both are True, no change needed

    processed_cells = 0
    # Iterate through all cells within the sheet's used range
    for row_idx in range(1, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)

            # Skip empty cells
            if cell.value is None or str(cell.value).strip() == "":
                continue

            value_str = str(cell.value).strip()
            # Determine strike status from cell font
            strike = bool(cell.font and cell.font.strike)
            processed_cells += 1

            # --- Skill Expression Processing (Specific Handling) ---
            if ">" in value_str:
                # Normalize expression part
                formatted_expr = value_str.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")
                ideal_agent_value = ""
                # Get corresponding Ideal Agent value if column was found
                if ideal_agent_col_idx:
                    # Check bounds before accessing cell
                    if ideal_agent_col_idx <= sheet.max_column:
                        ideal_cell = sheet.cell(row=row_idx, column=ideal_agent_col_idx)
                        ideal_agent_value = str(ideal_cell.value).strip() if ideal_cell.value else ""
                        # Normalize ideal part as well
                        ideal_agent_value = ideal_agent_value.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")
                    else:
                        # Log if ideal agent column index is out of bounds for this sheet
                        logging.warning(f"Ideal agent column {ideal_agent_col_idx} out of bounds for sheet '{sheet.title}' (max col: {sheet.max_column}) at row {row_idx}")

                # Create the concatenated key used for comparison and storage
                concatenated_value = f"{formatted_expr} {ideal_agent_value}".strip()
                if not concatenated_value:
                    # Skip if the key becomes empty after processing (e.g., only operators left)
                    continue

                # Update intermediate_data for skill_exprs directly (handles strike rule)
                skill_expr_dict = intermediate_data["skill_exprs"]
                if concatenated_value not in skill_expr_dict:
                    # Add new skill expression entry with all details
                    skill_expr_dict[concatenated_value] = {
                        "strike": strike,
                        "style_cell": cell,
                        "expr": formatted_expr,  # Store separate expression
                        "ideal": ideal_agent_value # Store separate ideal expression
                    }
                    logging.debug(f"Sheet '{sheet.title}' Row {cell.row}: Added new skill_expr '{concatenated_value}' with strike={strike}")
                elif skill_expr_dict[concatenated_value]["strike"] and not strike:
                    # Update existing entry if changing strike from True to False
                    skill_expr_dict[concatenated_value]["strike"] = False
                    skill_expr_dict[concatenated_value]["style_cell"] = cell # Use non-struck cell style
                    # Keep original expr/ideal values associated with the key
                    logging.debug(f"Sheet '{sheet.title}' Row {cell.row}: Updated skill_expr '{concatenated_value}' to strike=False")

                # Also extract individual skills from the expression part for the "Skills" sheet
                individual_skills = extract_skills(formatted_expr)
                for skill in individual_skills:
                     update_intermediate_generic(intermediate_data["skills"], skill, strike, cell)

            # --- VQ Processing ---
            # Check if it looks like a VQ name and NOT a skill expression
            elif (value_str.startswith("VQ_") or value_str.startswith("vq_") or ("VQ" in value_str)) and ">" not in value_str:
                 normalized_vq = value_str.replace(" ", "").replace('\u00A0', '')
                 update_intermediate_generic(intermediate_data["vqs"], normalized_vq, strike, cell)

            # --- VAG Processing (Only from specific sheet defined in config) ---
            elif "VAG_" in value_str and sheet.title == vag_sheet_name:
                 normalized_vag = value_str.replace(" ", "").replace('\u00A0', '')
                 update_intermediate_generic(intermediate_data["vags"], normalized_vag, strike, cell)

            # --- Simple Skill Check (Optional - Add if simple skills appear outside expressions) ---
            # Example: Check if it looks like a skill name and not other patterns
            # elif re.match(r'^[a-zA-Z0-9_]+$', value_str) and not any(x in value_str for x in ['>', 'VQ_', 'VAG_']):
            #      update_intermediate_generic(intermediate_data["skills"], value_str, strike, cell)

    logging.debug(f"Processed {processed_cells} non-empty cells in sheet '{sheet.title}'.")


def collect_routing_entities(
    workbook: openpyxl.workbook.Workbook,
    config: Dict[str, Any]
) -> Tuple[Dict[str, Set[str]], Dict[str, Dict[str, Dict[str, Any]]]]:
    """
    Processes workbook sheets using intermediate data structure, populates final
    output sheets based on resolved strike status (incl. separate expr/ideal for Skill Expr),
    and returns sheet data (non-struck only) for comparison AND the full intermediate data.
    Modifies the workbook object in place by deleting/adding sheets.

    Args:
        workbook: The openpyxl Workbook object (loaded from the working copy).
        config: The loaded configuration dictionary.

    Returns:
        Tuple containing:
        - sheet_data_for_comparison: Dict[str, Set[str]] - Sets of non-struck keys for comparison.
        - intermediate_data: Dict[str, Dict[str, Dict[str, Any]]] - Full collected data with details.
    """
    logging.info("Starting collection and processing of routing entities from workbook sheets.")

    # --- Intermediate Data Structure ---
    # Stores details for each unique item found across all sheets
    intermediate_data = {
        "vqs": {}, "skills": {}, "skill_exprs": {}, "vags": {}
    } # Format: {"Type": {"Name/Key": {"strike": bool, "style_cell": Cell, "expr"?: str, "ideal"?: str}}}

    # --- Phase 1: Process all sheets and populate intermediate_data ---
    for sheet in workbook.worksheets:
        process_sheet(sheet, intermediate_data, config)
    logging.info("Finished processing all sheets, resolved strikethrough status.")

    # --- Phase 2: Create/Populate Output Sheets from intermediate_data ---
    logging.info("Populating dedicated output sheets (VQ, Skills, Skill Expr, VAG)...")

    # Define structure and headers for the output sheets
    output_sheet_specs = {
        "Skill Expr": {"key": "skill_exprs", "headers": ["Expression", "Ideal Expression", "Concatenated Key", "HasStrikeThrough"]},
        "VQ": {"key": "vqs", "headers": ["VQ Name", "HasStrikeThrough"]},
        "VAG": {"key": "vags", "headers": ["VAG Name", "HasStrikeThrough"]},
        "Skills": {"key": "skills", "headers": ["Skill", "HasStrikeThrough"]}
    }

    # Define names for comparison sheets to ensure they are removed
    comparison_prefixes = {"Skill Expr": "Skill_exprs", "VQ": "Vqs", "VAG": "Vags", "Skills": "Skills"}
    # List all sheets to remove (output, comparison, metadata)
    sheets_to_remove = list(output_sheet_specs.keys()) + \
                       [f"{comparison_prefixes[t]} Comparison" for t in output_sheet_specs.keys()] + \
                       [METADATA_SHEET_NAME]

    # Remove old output/comparison/metadata sheets before creating new ones
    for sheet_name in sheets_to_remove:
         if sheet_name in workbook.sheetnames:
             try:
                 del workbook[sheet_name]
                 logging.debug(f"Removed existing sheet: {sheet_name}")
             except Exception as e:
                 logging.warning(f"Could not remove sheet '{sheet_name}': {e}")

    # Create and populate new output sheets
    for title, spec in output_sheet_specs.items():
        sheet = workbook.create_sheet(title=title)
        # Write Headers and make bold
        for col_idx, header in enumerate(spec["headers"], start=1):
             sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)

        row_num = 2 # Start writing data from row 2
        data_items = intermediate_data.get(spec["key"], {}) # Get data for this type

        # Sort items alphabetically by key for consistent output order
        for item_key, data in sorted(data_items.items()):
             if title == "Skill Expr":
                 # Populate the 4 columns for Skill Expr sheet
                 sheet.cell(row=row_num, column=1, value=data.get("expr", "")) # Expression
                 sheet.cell(row=row_num, column=2, value=data.get("ideal", "")) # Ideal Expression
                 cell_key = sheet.cell(row=row_num, column=3, value=item_key) # Concatenated Key
                 sheet.cell(row=row_num, column=4, value=str(data["strike"])) # HasStrikeThrough
                 # Copy style from the representative cell found during processing to the Key cell
                 copy_cell_style(data["style_cell"], cell_key)
             else:
                 # Standard 2-column population for VQ, VAG, Skills
                 cell_a = sheet.cell(row=row_num, column=1, value=item_key) # Name
                 sheet.cell(row=row_num, column=2, value=str(data["strike"])) # HasStrikeThrough
                 # Copy style from the representative cell
                 copy_cell_style(data["style_cell"], cell_a)
             row_num += 1
        logging.debug(f"Populated '{title}' sheet with {row_num - 2} items.")
    logging.info("Finished populating output sheets.")

    # --- Phase 3: Prepare data for comparison (only non-struck items' keys) ---
    # This data structure is used by write_comparison_sheet
    sheet_data_for_comparison = {
        key: {name for name, data in items.items() if not data["strike"]} # Set of keys where strike is False
        for key, items in intermediate_data.items()
    }

    logging.info("Prepared final sheet data for comparison (non-struck items only).")
    logging.debug(f"Comparison Data Summary: VQs={len(sheet_data_for_comparison['vqs'])}, Skills={len(sheet_data_for_comparison['skills'])}, SkillExprs={len(sheet_data_for_comparison['skill_exprs'])}, VAGs={len(sheet_data_for_comparison['vags'])}")

    # Return both the comparison keys and the full intermediate data with details
    return sheet_data_for_comparison, intermediate_data


# --- Comparison and Reporting ---
def write_comparison_sheet(
    workbook: openpyxl.workbook.Workbook,
    sheet_data: Dict[str, Set[str]],
    api_data: Dict[str, Dict[str, Any]], # API data structure varies by key
    intermediate_data: Dict[str, Dict[str, Dict[str, Any]]] # Full sheet data details
):
    """
    Compares sheet data (non-struck only) with API data and writes results
    to dedicated comparison sheets, including Expression/Ideal columns for Skill Exprs.

    Args:
        workbook: The openpyxl Workbook object to write to.
        sheet_data: Dict containing sets of non-struck item keys found in sheets.
        api_data: Dict containing items fetched from the API.
        intermediate_data: Dict containing detailed items found in sheets (used for 'new' items).
    """
    logging.info("Starting comparison and writing results to comparison sheets.")
    # Basic checks for empty data
    if not api_data:
        logging.warning("API data is empty or None, skipping comparison writing.")
        return
    if not sheet_data:
        logging.warning("Sheet data for comparison is empty or None, skipping comparison writing.")
        return

    # Map internal keys to sheet title prefixes used in comparison sheet names
    comparison_keys_map = {
        "vqs": "Vqs", "skills": "Skills",
        "skill_exprs": "Skill_exprs", "vags": "Vags"
    }

    for key, sheet_title_prefix in comparison_keys_map.items():
        logging.info(f"Generating comparison sheet for: {key}")
        comparison_sheet_title = f"{sheet_title_prefix} Comparison"
        # Sheet should have been removed in collect_routing_entities, but double-check
        if comparison_sheet_title in workbook.sheetnames:
             del workbook[comparison_sheet_title]
        sheet = workbook.create_sheet(title=comparison_sheet_title)

        # --- Prepare data for comparison ---
        sheet_items_non_struck = sheet_data.get(key, set()) # Keys of non-struck items from sheet
        api_items_dict = api_data.get(key, {}) # Dict of API items {key: details_or_id}
        api_items_keys = set(api_items_dict.keys()) # Keys of items found in API

        # Calculate differences based on the KEYS
        new_in_sheet = sheet_items_non_struck - api_items_keys
        missing_from_sheet_non_struck = api_items_keys - sheet_items_non_struck

        row_num = 2 # Start writing data from row 2

        # --- Set Headers and Column Widths based on key ---
        if key == "skill_exprs":
            headers = ["Concatenated Key", "Expression", "Ideal Expression", "ID (from API)", "Status"]
            # Adjust widths as needed for better viewing
            col_widths = [45, 45, 35, 20, 35]
        else:
            # Use the sheet_title_prefix (e.g., "Vqs", "Skills") as the first column header
            headers = [sheet_title_prefix, "ID (from API)", "Status"]
            col_widths = [45, 20, 35]

        # Write headers and apply formatting
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            # Set column width for better readability
            try:
                column_letter = openpyxl_cell_utils.get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].width = col_widths[col_idx-1]
            except IndexError: # Just in case col_widths definition has issues
                 pass # Ignore error if width definition is wrong


        # --- Write Data Rows ---
        # Write items New in Sheet (exist non-struck in sheet, but not in API)
        if new_in_sheet:
            logging.debug(f"'{key}' - Found {len(new_in_sheet)} items New in Sheet (Non-Struck).")
            # Sort items alphabetically for consistent report order
            for item_key in sorted(list(new_in_sheet)):
                if key == "skill_exprs":
                    # Lookup details from intermediate_data (originates from sheet processing)
                    item_details = intermediate_data['skill_exprs'].get(item_key, {})
                    sheet.cell(row=row_num, column=1, value=item_key) # Concatenated Key
                    sheet.cell(row=row_num, column=2, value=item_details.get('expr', '')) # Expression
                    sheet.cell(row=row_num, column=3, value=item_details.get('ideal', '')) # Ideal Expression
                    sheet.cell(row=row_num, column=4, value="N/A") # ID (Not applicable as it's not from API)
                    sheet.cell(row=row_num, column=5, value="New in Sheet (Non-Struck)") # Status
                else:
                    # Standard 3-column layout for VQ, Skill, VAG
                    sheet.cell(row=row_num, column=1, value=item_key) # Item Name
                    sheet.cell(row=row_num, column=2, value="N/A") # ID
                    sheet.cell(row=row_num, column=3, value="New in Sheet (Non-Struck)") # Status
                row_num += 1
        else:
             logging.debug(f"'{key}' - No items found only in the sheet (non-struck).")

        # Write items Missing from Sheet (exist in API, but not found non-struck in sheet)
        if missing_from_sheet_non_struck:
             logging.debug(f"'{key}' - Found {len(missing_from_sheet_non_struck)} items Missing from Sheet (or only Struck Out).")
             # Sort items alphabetically for consistent report order
             for item_key in sorted(list(missing_from_sheet_non_struck)):
                if key == "skill_exprs":
                    # Lookup details from api_data (originates from API)
                    # api_items_dict is api_data['skill_exprs'] here
                    api_details = api_items_dict.get(item_key, {}) # api_details is {'id': ..., 'expr': ..., 'ideal': ...}
                    sheet.cell(row=row_num, column=1, value=item_key) # Concatenated Key
                    sheet.cell(row=row_num, column=2, value=api_details.get('expr', '')) # Expression from API
                    sheet.cell(row=row_num, column=3, value=api_details.get('ideal', '')) # Ideal Expression from API
                    sheet.cell(row=row_num, column=4, value=api_details.get('id', 'ID Not Found')) # ID from API
                    sheet.cell(row=row_num, column=5, value="Missing in Sheet (or only Struck Out)") # Status
                else:
                    # Standard 3-column layout for VQ, Skill, VAG
                    # api_items_dict is api_data['vqs'], etc. here, value is just the ID string
                    api_id = api_items_dict.get(item_key, "ID Not Found") # api_id is just the string ID here
                    sheet.cell(row=row_num, column=1, value=item_key) # Item Name
                    sheet.cell(row=row_num, column=2, value=api_id) # ID from API
                    sheet.cell(row=row_num, column=3, value="Missing in Sheet (or only Struck Out)") # Status
                row_num += 1
        else:
            logging.debug(f"'{key}' - No items found only in the API (when compared to non-struck sheet items).")

        logging.info(f"Finished comparison sheet for: {key}")


# --- Main Orchestration ---
def main():
    """Main function to orchestrate the workbook processing and API comparison."""
    try:
        # Load configuration first
        config = load_configuration(CONFIG_FILE)
        source_file_path = config['source_file']
    except (FileNotFoundError, ValueError, configparser.Error) as config_e:
        logging.error(f"Halting execution due to configuration error: {config_e}")
        print(f"FATAL: Configuration Error - {config_e}. See {LOG_FILE} for details.")
        return # Stop execution

    # Determine output file path based on source file name
    base_name = os.path.splitext(source_file_path)[0]
    final_output_path = base_name + "_processed.xlsx"

    logging.info(f"--- Starting Processing Run ---")
    logging.info(f"Source Workbook: '{source_file_path}'")
    logging.info(f"Output Workbook: '{final_output_path}'")

    # Create a working copy of the source .xlsx file
    try:
        shutil.copyfile(source_file_path, final_output_path)
        logging.info(f"Copied source to '{final_output_path}' for processing.")
    except FileNotFoundError:
         logging.error(f"Source file not found: '{source_file_path}'. Cannot proceed.")
         print(f"FATAL: Source file '{source_file_path}' not found.")
         return
    except Exception as e:
        logging.error(f"Error copying file '{source_file_path}' to '{final_output_path}': {e}", exc_info=True)
        print(f"FATAL: Error copying source file. See {LOG_FILE} for details.")
        return

    workbook = None # Initialize workbook variable for finally block
    try:
        # --- Load the copied workbook ---
        logging.info(f"Loading workbook: {final_output_path}")
        workbook = openpyxl.load_workbook(final_output_path, read_only=False, data_only=False)

        # --- Step 1: Fetch API data AND Separate Max IDs ---
        # fetch_api_data now returns api_data, max_dn_id, max_ag_id
        api_data, max_dn_id, max_ag_id = fetch_api_data(config) # Unpack both IDs
        if not any(api_data.values()):
            logging.warning("API data fetch resulted in empty or partially empty datasets. Max IDs might be inaccurate.")

        # --- Step 2: Collect data from Excel sheets ---
        # This modifies the workbook by adding output sheets
        sheet_data_for_comparison, intermediate_data = collect_routing_entities(workbook, config)

        # --- Step 3: Perform comparison and write comparison sheets ---
        # This modifies the workbook by adding comparison sheets
        write_comparison_sheet(workbook, sheet_data_for_comparison, api_data, intermediate_data)

        # --- Step 4: Write Metadata (Separate Max IDs) ---
        # Create or get the Metadata sheet
        if METADATA_SHEET_NAME in workbook.sheetnames:
            metadata_sheet = workbook[METADATA_SHEET_NAME]
            logging.debug(f"Using existing '{METADATA_SHEET_NAME}' sheet.")
        else:
            metadata_sheet = workbook.create_sheet(title=METADATA_SHEET_NAME)
            logging.debug(f"Created new '{METADATA_SHEET_NAME}' sheet.")

        # Write labels and values for both Max IDs
        metadata_sheet[MAX_DN_ID_LABEL_CELL] = "Max DN API ID Found"
        metadata_sheet[MAX_DN_ID_LABEL_CELL].font = Font(bold=True)
        metadata_sheet[MAX_DN_ID_VALUE_CELL] = max_dn_id
        logging.info(f"Wrote Max DN API ID ({max_dn_id}) to '{METADATA_SHEET_NAME}' sheet, cell {MAX_DN_ID_VALUE_CELL}.")

        metadata_sheet[MAX_AG_ID_LABEL_CELL] = "Max AgentGroup API ID Found"
        metadata_sheet[MAX_AG_ID_LABEL_CELL].font = Font(bold=True)
        metadata_sheet[MAX_AG_ID_VALUE_CELL] = max_ag_id
        logging.info(f"Wrote Max AgentGroup API ID ({max_ag_id}) to '{METADATA_SHEET_NAME}' sheet, cell {MAX_AG_ID_VALUE_CELL}.")

        # --- Step 5: Save the processed workbook ---
        logging.info(f"Saving final processed workbook to: {final_output_path}")
        workbook.save(final_output_path)
        logging.info(f"Successfully saved processed workbook.")
        logging.info(f"--- Processing Complete ---")
        print(f"\nProcessing complete. Output saved to '{final_output_path}'")
        print(f"Log file saved to '{LOG_FILE}'")

    except Exception as e:
        # Catch any unexpected errors during processing
        logging.error(f"An unexpected error occurred during the main processing: {e}", exc_info=True)
        print(f"FATAL: An unexpected error occurred during processing. See {LOG_FILE} for details.")
    finally:
         # Ensure workbook is closed if it was opened, releasing file handle
         if workbook:
             try:
                 workbook.close()
                 logging.debug("Workbook closed.")
             except Exception as close_e:
                 logging.warning(f"Error closing workbook: {close_e}")


# --- Script Execution Guard ---
if __name__ == "__main__":
    # This block runs only when the script is executed directly
    main()
