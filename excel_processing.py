# -*- coding: utf-8 -*-
"""
Standalone Built-in Parser for processing an original source Excel file.

This module takes an original Excel workbook, identifies common routing entities
(VQs, Skills, VAGs, Skill Expressions) based on predefined logic and hints,
discards any struck-through items, performs cleaning, and outputs a new
workbook object with standardized sheets for each entity type.

This script is designed to be potentially user-editable for its parsing logic
if the default behavior needs adjustment for specific source Excel formats.
It aims to be self-contained for its parsing duties.
"""

import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import cell as openpyxl_cell_utils
import re
from typing import Dict, Any, Optional, Tuple, Set, List

# --- Logger for this module ---
# If run standalone, this will configure a basic logger.
# If imported by Flask app, it will use the app's logger config.
logger = logging.getLogger(__name__)
if not logger.hasHandlers(): # Avoid adding handlers if already configured by Flask
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(module)s:%(funcName)s] - %(message)s')

# --- Default Configuration Hints for the Built-in Parser ---
# Users can modify these directly if their source Excel has different conventions.
DEFAULT_IDEAL_AGENT_HEADER_TEXT = "Ideal Agent"
DEFAULT_IDEAL_AGENT_FALLBACK_CELL = "C2" # e.g., if header is not in row 1 of C or D
DEFAULT_VAG_EXTRACTION_SHEET_NAME = "Default Targeting- Group"
# Sheets to always skip when this parser processes an original Excel file
DEFAULT_SHEETS_TO_SKIP_PARSING = {
    "Instructions", "Summary", "Metadata",
    # Avoid re-processing sheets that might be named like its own output
    "VQs", "Skills", "Skill_Expressions", "VAGs"
}


# --- Internalized Utility Functions ---
def _extract_skills_from_expression(expression: str) -> list[str]:
    """
    Extracts potential skill names (alphanumeric + underscore) from a skill
    expression string. Looks for patterns like 'SkillName>5'.
    Internal to this module now.
    """
    if not isinstance(expression, str):
        return []
    # Regex finds words (alphanumeric or underscore) followed immediately by '>' and one or more digits
    skills = re.findall(r'\b([a-zA-Z0-9_]+)(?=>\d+)', expression)
    logger.debug(f"Extracted skills {skills} from expression '{expression}'")
    return skills

def _identify_ideal_agent_column(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    ideal_agent_header_text: str, # Expects header text string
    ideal_agent_fallback_cell: str # Expects fallback cell string
) -> Optional[int]:
    """
    Identifies the column index for the 'Ideal Agent' column based on hints.
    Searches header row 1 (specifically Columns C & D) first, then checks a fallback cell.
    Internal to this module.
    """
    logger.debug(f"Identifying '{ideal_agent_header_text}' column in sheet: {sheet.title}")
    # Check common header locations first (e.g., Columns C and D in row 1)
    for col_idx in [3, 4]:  # Column C=3, Column D=4
        if col_idx <= sheet.max_column:
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value and ideal_agent_header_text in str(cell_value):
                logger.debug(f"Found '{ideal_agent_header_text}' in header row 1 at column {col_idx}")
                return col_idx
    # Check fallback cell if not found
    try:
        col_str, row_str = openpyxl_cell_utils.coordinate_to_tuple(ideal_agent_fallback_cell)
        fallback_col_idx = openpyxl_cell_utils.column_index_from_string(col_str)
        fallback_row_idx = int(row_str)
        if fallback_row_idx <= sheet.max_row and fallback_col_idx <= sheet.max_column:
            cell_value_fallback = sheet.cell(row=fallback_row_idx, column=fallback_col_idx).value
            if cell_value_fallback and ideal_agent_header_text in str(cell_value_fallback):
                logger.debug(f"Found '{ideal_agent_header_text}' at fallback cell {ideal_agent_fallback_cell} (Col {fallback_col_idx})")
                return fallback_col_idx
    except Exception as e:
         logger.warning(f"Could not parse or check fallback cell '{ideal_agent_fallback_cell}': {e}")
    logger.debug(f"'{ideal_agent_header_text}' column not found in sheet: {sheet.title} using configured hints.")
    return None


# --- Main Parser Function ---
def parse_source_excel_to_standardized_workbook(
    source_workbook: openpyxl.workbook.Workbook,
    config_hints: Optional[Dict[str, Any]] = None
) -> openpyxl.workbook.Workbook:
    """
    Parses the original source Excel workbook to extract entities based on built-in logic.
    It discards struck-through items, cleans data, and creates a new
    workbook object with standardized output sheets for each entity type.
    This new workbook is intended to be the "_processed.xlsx" file.

    Args:
        source_workbook: The openpyxl.Workbook object of the original uploaded Excel.
                         Expected to be loaded with style information to detect strikethrough.
        config_hints: An optional dictionary containing hints from the application's
                      config.ini (e.g., 'ideal_agent_header_text'). Uses internal
                      defaults if not provided.

    Returns:
        A new openpyxl.Workbook object containing the parsed and standardized entity sheets.
    """
    logger.info("Starting built-in parsing of source workbook to create standardized entity sheets...")

    # Use provided config hints or defaults
    cfg = config_hints if config_hints is not None else {}
    ideal_agent_header = cfg.get('ideal_agent_header_text', DEFAULT_IDEAL_AGENT_HEADER_TEXT)
    ideal_agent_fallback = cfg.get('ideal_agent_fallback_cell', DEFAULT_IDEAL_AGENT_FALLBACK_CELL)
    vag_sheet_name_hint = cfg.get('vag_extraction_sheet', DEFAULT_VAG_EXTRACTION_SHEET_NAME)
    sheets_to_skip = DEFAULT_SHEETS_TO_SKIP_PARSING.copy() # Use a copy

    # Intermediate storage for unique, non-struck entities
    parsed_data = {
        "VQs": set(),
        "Skills": set(),
        "Skill_Expressions": [], # List of dicts for structure
        "VAGs": set()
    }

    # --- Iterate through sheets and cells of the source workbook ---
    for sheet in source_workbook.worksheets:
        if sheet.title in sheets_to_skip:
            logger.debug(f"Skipping sheet during parsing: {sheet.title} (in predefined skip list)")
            continue

        logger.info(f"Built-in parser processing sheet: {sheet.title}")
        # --- MODIFICATION START: Pass individual config hints ---
        ideal_agent_col_idx = _identify_ideal_agent_column(
            sheet,
            ideal_agent_header,
            ideal_agent_fallback
        )
        # --- MODIFICATION END ---

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None: # Skip truly empty cells
                    continue
                
                value_str_raw = str(cell.value) # Get raw value before stripping for logging
                value_str_stripped = value_str_raw.strip()

                if not value_str_stripped: # Skip cells that are empty after stripping
                    continue

                # --- Check for strikethrough: if struck, discard this item ---
                if cell.font and cell.font.strike:
                    logger.debug(f"Skipping struck-through cell {cell.coordinate}: '{value_str_raw}'")
                    continue # Skip this cell's value entirely

                # Identify VQs (starts with "VQ_", not a skill expression)
                if (value_str_stripped.lower().startswith("vq_") or "vq" in value_str_stripped.lower()) and ">" not in value_str_stripped:
                    cleaned_vq = value_str_stripped.replace(" ", "").replace('\u00A0', '')
                    if cleaned_vq: # Ensure not empty after cleaning
                        parsed_data["VQs"].add(cleaned_vq)
                        logger.debug(f"Parser found VQ: {cleaned_vq} from {cell.coordinate}")

                # Identify Skill Expressions (contains ">")
                elif ">" in value_str_stripped:
                    raw_expression = value_str_stripped # Already stripped
                    ideal_expression_str = ""
                    if ideal_agent_col_idx and ideal_agent_col_idx <= sheet.max_column:
                        ideal_cell = sheet.cell(row=row_idx, column=ideal_agent_col_idx)
                        if ideal_cell.value is not None:
                            if not (ideal_cell.font and ideal_cell.font.strike):
                                ideal_expression_str = str(ideal_cell.value).strip()
                            else:
                                logger.debug(f"Ideal Agent cell {ideal_cell.coordinate} is struck-through for expression '{raw_expression}'. Ignoring ideal part.")

                    cleaned_expression = raw_expression.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")
                    cleaned_ideal = ideal_expression_str.replace(" ", "").replace('\u00A0', '').replace("|", " | ").replace("&", " & ")
                    concatenated_key = cleaned_expression
                    if cleaned_ideal: concatenated_key = f"{cleaned_expression} {cleaned_ideal}".strip()
                    extracted_skills_list_for_this_expr = []
                    if cleaned_expression:
                        skills_from_expr = _extract_skills_from_expression(cleaned_expression)
                        for skill in skills_from_expr:
                            cleaned_skill = skill.replace(" ", "").replace('\u00A0', '')
                            if cleaned_skill:
                                parsed_data["Skills"].add(cleaned_skill)
                                extracted_skills_list_for_this_expr.append(cleaned_skill)
                    if cleaned_expression:
                        parsed_data["Skill_Expressions"].append({
                            "Original Expression": cleaned_expression, "Ideal Expression": cleaned_ideal,
                            "Concatenated Key": concatenated_key,
                            "Extracted_Skills_List_String": ", ".join(extracted_skills_list_for_this_expr)
                        })
                        logger.debug(f"Parser found Skill Expression: {concatenated_key} from {cell.coordinate}")

                # Identify VAGs (starts with "VAG_", specific sheet only as per hint)
                elif value_str_stripped.startswith("VAG_") and sheet.title == vag_sheet_name_hint:
                    cleaned_vag = value_str_stripped.replace(" ", "").replace('\u00A0', '')
                    if cleaned_vag:
                        parsed_data["VAGs"].add(cleaned_vag)
                        logger.debug(f"Parser found VAG: {cleaned_vag} from {cell.coordinate}")
                
    logger.info("Built-in parser finished initial data extraction from source workbook.")

    # --- Create a new workbook for the processed output ---
    output_workbook = openpyxl.Workbook()
    if "Sheet" in output_workbook.sheetnames and len(output_workbook.sheetnames) == 1:
        try: output_workbook.remove(output_workbook.active)
        except Exception as e_rm_sheet: logger.warning(f"Could not remove default sheet: {e_rm_sheet}")

    bold_font = Font(bold=True)
    if parsed_data["VQs"]:
        vq_sheet = output_workbook.create_sheet("VQs")
        vq_sheet.cell(row=1, column=1, value="VQ Name").font = bold_font
        for i, vq_name in enumerate(sorted(list(parsed_data["VQs"])), start=2): vq_sheet.cell(row=i, column=1, value=vq_name)
        logger.info(f"Created 'VQs' output sheet with {len(parsed_data['VQs'])} items.")
    if parsed_data["Skills"]:
        skill_sheet = output_workbook.create_sheet("Skills")
        skill_sheet.cell(row=1, column=1, value="Skill Name").font = bold_font
        for i, skill_name in enumerate(sorted(list(parsed_data["Skills"])), start=2): skill_sheet.cell(row=i, column=1, value=skill_name)
        logger.info(f"Created 'Skills' output sheet with {len(parsed_data['Skills'])} items.")
    if parsed_data["VAGs"]:
        vag_sheet = output_workbook.create_sheet("VAGs")
        vag_sheet.cell(row=1, column=1, value="VAG Name").font = bold_font
        for i, vag_name in enumerate(sorted(list(parsed_data["VAGs"])), start=2): vag_sheet.cell(row=i, column=1, value=vag_name)
        logger.info(f"Created 'VAGs' sheet with {len(parsed_data['VAGs'])} items.")
    if parsed_data["Skill_Expressions"]:
        se_sheet = output_workbook.create_sheet("Skill_Expressions")
        se_headers = ["Original Expression", "Ideal Expression", "Concatenated Key", "Extracted_Skills_List_String"]
        for col_idx, header in enumerate(se_headers, start=1): se_sheet.cell(row=1, column=col_idx, value=header).font = bold_font
        sorted_skill_expressions = sorted(parsed_data["Skill_Expressions"], key=lambda x: x.get("Concatenated Key", ""))
        for row_idx, se_data in enumerate(sorted_skill_expressions, start=2):
            se_sheet.cell(row=row_idx, column=1, value=se_data.get("Original Expression"))
            se_sheet.cell(row=row_idx, column=2, value=se_data.get("Ideal Expression"))
            se_sheet.cell(row=row_idx, column=3, value=se_data.get("Concatenated Key"))
            se_sheet.cell(row=row_idx, column=4, value=se_data.get("Extracted_Skills_List_String"))
        logger.info(f"Created 'Skill_Expressions_Output' sheet with {len(parsed_data['Skill_Expressions'])} items.")

    logger.info("Built-in parser finished creating standardized output workbook object.")
    return output_workbook


# Example usage if run as a standalone script (for testing the parser)
if __name__ == '__main__': # pragma: no cover
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - [%(module)s:%(funcName)s] - %(message)s')
    logger.info("Testing excel_processing.py standalone parser.")
    test_wb_path = "test_source_excel.xlsx"
    wb = openpyxl.Workbook()
    sheet1 = wb.active; sheet1.title = "Sheet1_VQs_Skills"
    sheet1['A1'] = "VQ_SALES_EN"; sheet1['A2'] = "VQ_SUPPORT_ES"; sheet1['A2'].font = Font(strike=True)
    sheet1['B1'] = "SkillA>5 & SkillB>3"; sheet1['C1'] = "IdealA>0"
    sheet1['B2'] = "SkillC>2"; sheet1['A3'] = "VQ_Billing"; sheet_1['A4'] = "  VQ_ espa\u00A0ce  "
    sheet2 = wb.create_sheet("Default Targeting- Group")
    sheet2['A1'] = "VAG_Tier1_Support"; sheet2['A2'] = "VAG_Sales_VIP"; sheet2['A2'].font = Font(strike=True)
    wb.save(test_wb_path)
    logger.info(f"Created dummy test workbook: {test_wb_path}")
    loaded_test_wb = openpyxl.load_workbook(test_wb_path, data_only=False, read_only=False)
    test_config_hints = {
        'ideal_agent_header_text': 'Ideal Agent',
        'ideal_agent_fallback_cell': 'C1',
        'vag_extraction_sheet': 'Default Targeting- Group'
    }
    processed_wb = parse_source_excel_to_standardized_workbook(loaded_test_wb, test_config_hints)
    loaded_test_wb.close()
    processed_output_path = "test_source_excel_PARSED.xlsx"
    processed_wb.save(processed_output_path)
    logger.info(f"Saved parsed output to: {processed_output_path}")
    verify_wb = openpyxl.load_workbook(processed_output_path)
    print("\nGenerated Sheets:");
    for sheetname in verify_wb.sheetnames:
        print(f"- {sheetname}"); ws = verify_wb[sheetname]
        for row in ws.iter_rows(values_only=True): print(f"  {row}")
    verify_wb.close()
    print(f"\nTest complete. Check '{processed_output_path}'.")

