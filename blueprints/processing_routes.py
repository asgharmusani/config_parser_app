# -*- coding: utf-8 -*-
"""
Flask Blueprint for handling backend processing tasks and API endpoints.

Supports a two-stage workflow:
1. Built-in parsing of an original Excel file to create a standardized "_processed.xlsx" file.
2. Comparison of data from a "_processed.xlsx" file against APIs using a rule template.
"""

import os
import json
import logging
import re
import shutil
import openpyxl
import datetime # For timestamped filenames
from flask import (
    Blueprint, request, jsonify, current_app, abort, flash, redirect, url_for
)
from werkzeug.utils import secure_filename
from openpyxl.styles import Font
from typing import Dict, Any, Optional, Tuple, Set, List

# Import utility functions and constants
try:
    from utils import IdGenerator, replace_placeholders, read_comparison_data
    TEMPLATE_DIR = './config_templates/'
    EXCEL_RULE_TEMPLATE_DIR = './excel_rule_templates/'
except ImportError as e:
    logging.error(f"Failed to import required functions/constants for processing_routes: {e}")
    class IdGenerator:
        def __init__(self, *args, **kwargs): pass
        def get_next_dn_id(self): return 0
        def get_next_ag_id(self): return 0
    def replace_placeholders(template_data, row_data, current_row_next_id=None): return template_data
    def read_comparison_data(filename: str) -> bool: return False
    TEMPLATE_DIR = './config_templates/'
    EXCEL_RULE_TEMPLATE_DIR = './excel_rule_templates/'

# Import functions from the refactored processing modules
try:
    from config import save_config
    from excel_processing import parse_source_excel_to_standardized_workbook as built_in_parse_source_excel
    from api_fetching import fetch_and_process_api_data_for_entity
    from comparison_logic import write_comparison_sheets
    METADATA_SHEET_NAME = "Metadata"
    MAX_DN_ID_LABEL_CELL = "A1"; MAX_DN_ID_VALUE_CELL = "B1"
    MAX_AG_ID_LABEL_CELL = "A2"; MAX_AG_ID_VALUE_CELL = "B2"
    DN_SHEETS = {"Vqs Comparison", "VQs"}
    AGENT_GROUP_SHEETS = {
        "Skills Comparison", "Vags Comparison", "Skill_exprs Comparison",
        "Skills", "VAGs", "Skill_Expressions"
    }
except ImportError as e:
     logging.critical(f"CRITICAL: Failed to import core processing functions: {e}. Processing endpoints will fail.", exc_info=True)
     def save_config(p, s): raise NotImplementedError("save_config not imported")
     def built_in_parse_source_excel(wb, cfg): raise NotImplementedError("built_in_parse_source_excel not imported")
     def fetch_and_process_api_data_for_entity(u, en, r, c): return ({}, 0)
     def write_comparison_sheets(w, s, a, i): raise NotImplementedError("write_comparison_sheets not imported")
     METADATA_SHEET_NAME = "Metadata"; MAX_DN_ID_LABEL_CELL = "A1"; MAX_DN_ID_VALUE_CELL = "B1"; MAX_AG_ID_LABEL_CELL = "A2"; MAX_AG_ID_VALUE_CELL = "B2"
     DN_SHEETS = set(); AGENT_GROUP_SHEETS = set()

# --- Constants ---
UPLOAD_FOLDER = './uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

# --- Logging ---
logger = logging.getLogger(__name__)

# --- Blueprint Definition ---
processing_bp = Blueprint('processing', __name__)

# --- Helper Functions ---
def allowed_file(filename: str) -> bool:
    """Checks if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# --- API Routes ---

@processing_bp.route('/upload-original-file', methods=['POST'])
def upload_original_file():
    """
    DEPRECATED (or for standalone upload if needed).
    The /run-comparison route now handles direct upload for new file processing.
    Handles upload of the original source Excel file.
    Saves the file and stores its path for later processing.
    """
    logger.info("Received request to /upload-original-file.")
    # This route might still be useful if you want a separate upload step before processing.
    # For the current UI flow of "Process New File", it's combined into /run-comparison.
    if 'sourceExcelFile' not in request.files:
        logger.warning("File upload request missing 'sourceExcelFile' part.")
        return jsonify({"error": "No file part in the request."}), 400

    file = request.files['sourceExcelFile']
    if file.filename == '':
        logger.warning("File upload request received with no selected file.")
        return jsonify({"error": "No selected file."}), 400

    if file and allowed_file(file.filename):
        if not os.path.exists(UPLOAD_FOLDER):
            try: os.makedirs(UPLOAD_FOLDER); logger.info(f"Created upload directory: {UPLOAD_FOLDER}")
            except OSError as e: logger.error(f"Could not create upload directory {UPLOAD_FOLDER}: {e}"); return jsonify({"error": f"Server error: Could not create upload directory."}), 500

        original_filename = secure_filename(file.filename)
        original_filepath = os.path.join(UPLOAD_FOLDER, original_filename)

        try:
            file.save(original_filepath)
            logger.info(f"Uploaded original file saved to: {original_filepath}")
            # Store path for potential later use if run_comparison is called separately
            current_app.config['LAST_UPLOADED_ORIGINAL_FILE_PATH'] = original_filepath
            current_app.config['LAST_UPLOADED_ORIGINAL_FILENAME'] = original_filename
            return jsonify({
                "message": f"File '{original_filename}' uploaded successfully. Ready to process.",
                "original_filename": original_filename
                }), 200
        except Exception as e:
            logger.error(f"Error saving uploaded file: {e}", exc_info=True)
            return jsonify({"error": f"An error occurred saving the uploaded file: {e}"}), 500
    else:
        logger.warning(f"Invalid file type uploaded: {file.filename}")
        return jsonify({"error": "Invalid file type. Please upload an .xlsx file."}), 400


@processing_bp.route('/run-comparison', methods=['POST'])
def run_comparison():
    """
    Orchestrates the "Process New File" workflow.
    Receives an original Excel file via FormData, an Excel Rule Template name (optional if not comparing),
    and a flag to perform comparison.
    """
    logger.info("Request received for /run-comparison (Process New File workflow)")

    # --- MODIFICATION START: Handle direct file upload within this route ---
    if 'sourceExcelFile' not in request.files:
        logger.warning("/run-comparison: Missing 'sourceExcelFile' in request files.")
        return jsonify({"error": "No sourceExcelFile part in the request."}), 400
    
    file = request.files['sourceExcelFile']
    if file.filename == '':
        logger.warning("/run-comparison: No selected source file.")
        return jsonify({"error": "No selected source file."}), 400

    if not (file and allowed_file(file.filename)):
        logger.warning(f"/run-comparison: Invalid file type uploaded: {file.filename}")
        return jsonify({"error": "Invalid file type. Please upload an .xlsx file."}), 400

    original_filename = secure_filename(file.filename)
    # Save the uploaded file to a temporary path or directly to its final original path
    # For simplicity, save it directly to where it would be processed from.
    if not os.path.exists(UPLOAD_FOLDER):
        try: os.makedirs(UPLOAD_FOLDER)
        except OSError as e: logger.error(f"Could not create upload dir: {e}"); return jsonify({"error": "Server error creating upload directory."}), 500
    
    original_filepath = os.path.join(UPLOAD_FOLDER, original_filename)
    try:
        file.save(original_filepath)
        logger.info(f"/run-comparison: Uploaded original file saved to: {original_filepath}")
    except Exception as e:
        logger.error(f"/run-comparison: Error saving uploaded file: {e}", exc_info=True)
        return jsonify({"error": f"An error occurred saving the uploaded file: {e}"}), 500
    # --- MODIFICATION END ---

    excel_rule_template_name = request.form.get('excelRuleTemplateName')
    perform_comparison_str = request.form.get('perform_comparison', 'false')
    perform_comparison = perform_comparison_str.lower() == 'true'

    if perform_comparison and not excel_rule_template_name:
        logger.warning("/run-comparison: (perform_comparison=true) missing 'excelRuleTemplateName'.")
        # Clean up uploaded file if rule is missing for comparison
        if os.path.exists(original_filepath): os.remove(original_filepath)
        return jsonify({"error": "Excel rule template name is required when performing comparison."}), 400

    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    processed_filename = f"{os.path.splitext(original_filename)[0]}_{timestamp}_processed.xlsx"
    processed_filepath = os.path.join(UPLOAD_FOLDER, processed_filename)

    logger.info(f"Processing new file: '{original_filename}', Rule: '{excel_rule_template_name if perform_comparison else 'Built-in Parser Only'}', Compare: {perform_comparison}, Output: '{processed_filepath}'")
    app_config_settings = current_app.config.get('APP_SETTINGS', {})
    rule_template_json = None

    if perform_comparison and excel_rule_template_name:
        rule_template_path = os.path.join(EXCEL_RULE_TEMPLATE_DIR, excel_rule_template_name)
        if not os.path.exists(rule_template_path):
            logger.error(f"Excel rule template file not found: {rule_template_path}")
            if os.path.exists(original_filepath): os.remove(original_filepath) # Cleanup
            return jsonify({"error": f"Excel rule template '{excel_rule_template_name}' not found."}), 404
        try:
            with open(rule_template_path, 'r', encoding='utf-8') as f: rule_template_json = json.load(f)
            logger.info(f"Loaded Excel rule template: {excel_rule_template_name}")
        except Exception as e:
            logger.error(f"Error loading/parsing Excel rule template '{excel_rule_template_name}': {e}", exc_info=True)
            if os.path.exists(original_filepath): os.remove(original_filepath) # Cleanup
            return jsonify({"error": f"Could not load/parse Excel rule template: {e}"}), 500

    output_workbook = None
    try:
        source_workbook = openpyxl.load_workbook(original_filepath, read_only=False, data_only=False)
        parsed_workbook_object = built_in_parse_source_excel(source_workbook, app_config_settings)
        source_workbook.close()
        logger.info(f"Built-in parser finished processing '{original_filename}'.")
        output_workbook = parsed_workbook_object

        overall_max_dn_id = 0
        overall_max_ag_id = 0

        if perform_comparison:
            logger.info("Proceeding with API comparison.")
            api_data_for_comparison = {}
            if rule_template_json and "Entities" in rule_template_json:
                for entity_rule in rule_template_json["Entities"]:
                    if not entity_rule.get("enabled", True): continue
                    entity_name = entity_rule["name"]
                    api_url = entity_rule.get("comparisonApiUrl")
                    id_pool_type = entity_rule.get("idPoolType")
                    if api_url:
                        processed_data, max_id_this_api = fetch_and_process_api_data_for_entity(
                            api_url, entity_name, entity_rule, app_config_settings
                        )
                        api_data_for_comparison[entity_name] = processed_data
                        if id_pool_type == 'dn': overall_max_dn_id = max(overall_max_dn_id, max_id_this_api)
                        elif id_pool_type == 'agent_group': overall_max_ag_id = max(overall_max_ag_id, max_id_this_api)
                        elif max_id_this_api > 0: logger.warning(f"Max ID {max_id_this_api} for entity '{entity_name}' from API '{api_url}' has no recognized 'idPoolType'.")
                    else:
                        api_data_for_comparison[entity_name] = {}
            logger.info(f"Aggregated Max IDs from API calls: DN={overall_max_dn_id}, AG={overall_max_ag_id}")
            
            # Prepare data for comparison sheets from the parsed_workbook_object's sheets
            temp_sheet_data_for_comp = {}
            temp_intermediate_data = {} # For write_comparison_sheets to get details for "New in Sheet"
            
            # Iterate through sheets in the workbook created by built_in_parse_source_excel
            for sheet_name_in_parsed_wb in output_workbook.sheetnames:
                if sheet_name_in_parsed_wb == METADATA_SHEET_NAME: continue # Skip metadata if it somehow exists

                # The sheet_name_in_parsed_wb is an entity name (e.g., "VQs", "Skills")
                # We need to find the corresponding rule to know the primaryKeyColumnExcel
                pk_col_excel_for_entity = sheet_name_in_parsed_wb # Default to sheet name itself
                if rule_template_json: # Only needed if comparing, to find primaryKeyColumnExcel
                    for rule_def in rule_template_json.get("Entities", []):
                        if rule_def.get("name") == sheet_name_in_parsed_wb: # Match rule name with sheet name
                            pk_col_excel_for_entity = rule_def.get("primaryKeyColumnExcel", sheet_name_in_parsed_wb)
                            break
                
                sheet_obj = output_workbook[sheet_name_in_parsed_wb]
                headers = [cell.value for cell in sheet_obj[1]]
                headers = [str(h).strip() for h in headers if h is not None]
                if not headers or pk_col_excel_for_entity not in headers:
                    logger.warning(f"Cannot determine primary key column '{pk_col_excel_for_entity}' or headers for sheet '{sheet_name_in_parsed_wb}' in parsed workbook. Skipping for comparison prep.")
                    continue

                temp_sheet_data_for_comp[sheet_name_in_parsed_wb] = set()
                temp_intermediate_data[sheet_name_in_parsed_wb] = {}
                for row_cells in sheet_obj.iter_rows(min_row=2, values_only=False): # Get cell objects for style
                    row_data = {headers[i]: cell.value for i, cell in enumerate(row_cells) if i < len(headers)}
                    item_key = str(row_data.get(pk_col_excel_for_entity, ''))
                    if not item_key: continue
                    
                    # Built-in parser already removed struck-through items.
                    # For comparison logic, we assume 'strike' is false for all items from these sheets.
                    temp_sheet_data_for_comp[sheet_name_in_parsed_wb].add(item_key)
                    
                    current_item_details = row_data.copy()
                    current_item_details['strike'] = False # All items from parser are non-struck
                    # Store source info (though it's from the parsed sheet, not original cell directly)
                    current_item_details['_source_sheet_title_'] = sheet_name_in_parsed_wb
                    # The actual coordinate is from the parsed sheet, not the *very original* Excel.
                    # If styling is needed, it should be applied by the built-in parser.
                    current_item_details['_source_cell_coordinate_'] = row_cells[headers.index(pk_col_excel_for_entity)].coordinate if pk_col_excel_for_entity in headers else "N/A"
                    temp_intermediate_data[sheet_name_in_parsed_wb][item_key] = current_item_details

            write_comparison_sheets(
                output_workbook, temp_sheet_data_for_comp, api_data_for_comparison, temp_intermediate_data
            )

            # Write Metadata sheet with aggregated Max IDs
            if METADATA_SHEET_NAME in output_workbook.sheetnames: del output_workbook[METADATA_SHEET_NAME]
            metadata_sheet = output_workbook.create_sheet(title=METADATA_SHEET_NAME)
            metadata_sheet[MAX_DN_ID_LABEL_CELL] = "Max DN API ID Found"; metadata_sheet[MAX_DN_ID_LABEL_CELL].font = Font(bold=True)
            metadata_sheet[MAX_DN_ID_VALUE_CELL] = overall_max_dn_id
            metadata_sheet[MAX_AG_ID_LABEL_CELL] = "Max AgentGroup API ID Found"; metadata_sheet[MAX_AG_ID_LABEL_CELL].font = Font(bold=True)
            metadata_sheet[MAX_AG_ID_VALUE_CELL] = overall_max_ag_id
            logging.info(f"Wrote Aggregated Max IDs (DN:{overall_max_dn_id}, AG:{overall_max_ag_id}) to '{METADATA_SHEET_NAME}'.")

        # Save the final workbook (either just parsed or parsed+compared)
        output_workbook.save(processed_filepath)
        logger.info(f"Successfully saved final processed workbook to: {processed_filepath}")

        if perform_comparison:
            logger.info("Reloading application data cache from processed file (after comparison)...")
            current_app.config['EXCEL_DATA'] = {}; current_app.config['EXCEL_FILENAME'] = None; current_app.config['COMPARISON_SHEETS'] = []; current_app.config['SHEET_HEADERS'] = {}; current_app.config['MAX_DN_ID'] = 0; current_app.config['MAX_AG_ID'] = 0
            if read_comparison_data(processed_filepath):
                 logger.info("Application cache updated successfully.")
                 first_sheet = current_app.config.get('COMPARISON_SHEETS', [None])[0]
                 return jsonify({
                     "message": f"File '{original_filename}' processed and compared successfully using rule '{excel_rule_template_name}'.",
                     "processed_file": processed_filename,
                     "redirect_url": url_for('ui.view_comparison', comparison_type=first_sheet) if first_sheet else url_for('ui.upload_config_page')
                     }), 200
            else:
                 logger.error("Failed to reload data cache after processing and comparison.")
                 return jsonify({"error": f"File '{original_filename}' processed and compared, but failed to reload data into UI cache. Check logs."}), 500
        else: # "Parse Only" was successful
            return jsonify({
                "message": f"File '{original_filename}' parsed successfully using built-in parser. Output: {processed_filename}",
                "processed_file": processed_filename
                }), 200

    except Exception as proc_err:
        logger.error(f"Error during 'run_comparison' for '{original_filename}': {proc_err}", exc_info=True)
        return jsonify({"error": f"Error processing file '{original_filename}': {proc_err}"}), 500
    finally:
        if output_workbook:
            try: output_workbook.close()
            except Exception as close_e: logger.warning(f"Error closing output workbook: {close_e}")
        # Clean up the initially saved original file after processing
        if os.path.exists(original_filepath):
            try:
                os.remove(original_filepath)
                logger.info(f"Removed temporary original uploaded file: {original_filepath}")
            except OSError as rm_err:
                logger.warning(f"Could not remove temp original upload file {original_filepath}: {rm_err}")


@processing_bp.route('/load-processed-file', methods=['POST'])
def load_processed_file():
    """
    Loads an existing *_processed.xlsx file.
    If 'perform_comparison' is true, it re-runs API comparisons based on the
    selected rule template and updates the comparison sheets in that file.
    Then, loads data into the app cache for viewing.
    """
    logger.info("Request received to load/compare existing processed file.")
    request_data = request.get_json()
    if not request_data or 'filename' not in request_data:
        return jsonify({"error": "Filename not provided."}), 400

    processed_filename = secure_filename(request_data['filename'])
    excel_rule_template_name = request_data.get('excelRuleTemplateName') # This is the comparison rule
    perform_comparison_str = str(request_data.get('perform_comparison', 'false')).lower()
    perform_comparison = perform_comparison_str == 'true'

    if perform_comparison and not excel_rule_template_name:
        return jsonify({"error": "Comparison rule template name is required when performing comparison."}), 400

    processed_filepath = os.path.join(UPLOAD_FOLDER, processed_filename)
    logger.info(f"Loading/Comparing processed file: '{processed_filepath}', Rule: '{excel_rule_template_name if perform_comparison else 'None (Load Only)'}', Compare: {perform_comparison}")

    if not os.path.exists(processed_filepath):
        logger.error(f"Processed file not found: {processed_filepath}")
        return jsonify({"error": f"File '{processed_filename}' not found in uploads directory."}), 404

    app_config_settings = current_app.config.get('APP_SETTINGS', {})
    rule_template_json = None
    if perform_comparison: # Load rule template only if comparing
        rule_template_path = os.path.join(EXCEL_RULE_TEMPLATE_DIR, excel_rule_template_name)
        if not os.path.exists(rule_template_path):
            return jsonify({"error": f"Comparison rule template '{excel_rule_template_name}' not found."}), 404
        try:
            with open(rule_template_path, 'r', encoding='utf-8') as f: rule_template_json = json.load(f)
        except Exception as e:
            return jsonify({"error": f"Could not load/parse comparison rule template: {e}"}), 500

    output_workbook = None # Initialize for finally block
    try:
        # Clear existing cache before loading/re-processing
        current_app.config['EXCEL_DATA'] = {}; current_app.config['EXCEL_FILENAME'] = None; current_app.config['COMPARISON_SHEETS'] = []; current_app.config['SHEET_HEADERS'] = {}; current_app.config['MAX_DN_ID'] = 0; current_app.config['MAX_AG_ID'] = 0

        if not perform_comparison:
            # "Load Only" mode: Just read the file into cache
            if read_comparison_data(processed_filepath): # utils.read_comparison_data
                logger.info(f"Successfully loaded data from '{processed_filename}' into cache (Load Only).")
                first_sheet_to_view_list = current_app.config.get('COMPARISON_SHEETS', []) # Prefers comparison sheets
                if not first_sheet_to_view_list and current_app.config.get('EXCEL_DATA'): # Fallback to first data sheet
                    first_sheet_to_view_list = list(current_app.config.get('EXCEL_DATA').keys())
                
                first_sheet_to_view = first_sheet_to_view_list[0] if first_sheet_to_view_list else None
                redirect_url = url_for('ui.view_comparison', comparison_type=first_sheet_to_view) if first_sheet_to_view else url_for('ui.upload_config_page')

                return jsonify({
                    "message": f"Successfully loaded data from '{processed_filename}'.",
                    "redirect_url": redirect_url,
                    "redirect_url_for_view_only": redirect_url # For JS logic
                }), 200
            else:
                return jsonify({"error": f"Failed to read data from '{processed_filename}'. Check logs."}), 500

        # "Load and Compare" mode for an existing processed file
        if not read_comparison_data(processed_filepath): # This also loads Max IDs from its Metadata
            return jsonify({"error": f"Failed to initially read '{processed_filename}' for comparison. Check logs."}), 500
        
        loaded_excel_data = current_app.config.get('EXCEL_DATA', {})
        sheet_data_for_comparison_recomp = {}
        intermediate_data_recomp = {}

        for entity_rule in rule_template_json.get("Entities", []):
            if not entity_rule.get("enabled", True): continue
            entity_name = entity_rule["name"]
            source_sheet_name_from_rule = entity_rule.get("sourceSheetName", entity_name)
            primary_key_col_excel = entity_rule.get("primaryKeyColumnExcel")

            if source_sheet_name_from_rule not in loaded_excel_data:
                logger.warning(f"For re-compare, entity '{entity_name}': source sheet '{source_sheet_name_from_rule}' not found. Skipping.")
                sheet_data_for_comparison_recomp[entity_name] = set(); intermediate_data_recomp[entity_name] = {}; continue
            if not primary_key_col_excel:
                headers_for_source_sheet = current_app.config['SHEET_HEADERS'].get(source_sheet_name_from_rule)
                if headers_for_source_sheet: primary_key_col_excel = headers_for_source_sheet[0]
                else: logger.error(f"Cannot determine pk col for entity '{entity_name}'. Skipping."); sheet_data_for_comparison_recomp[entity_name] = set(); intermediate_data_recomp[entity_name] = {}; continue
            
            sheet_data_for_comparison_recomp[entity_name] = set()
            intermediate_data_recomp[entity_name] = {}
            for row_dict in loaded_excel_data[source_sheet_name_from_rule]:
                item_key = str(row_dict.get(primary_key_col_excel, ''));
                if not item_key: continue
                is_struck = str(row_dict.get("StrikeStatus", "false")).lower() == "true"
                if not is_struck: sheet_data_for_comparison_recomp[entity_name].add(item_key)
                temp_row_dict = row_dict.copy(); temp_row_dict['strike'] = is_struck
                intermediate_data_recomp[entity_name][item_key] = temp_row_dict

        api_data_for_comparison = {}; overall_max_dn_id_recomp = 0; overall_max_ag_id_recomp = 0
        if rule_template_json and "Entities" in rule_template_json:
            for entity_rule in rule_template_json["Entities"]:
                if not entity_rule.get("enabled", True): continue
                entity_name = entity_rule["name"]; api_url = entity_rule.get("comparisonApiUrl"); id_pool = entity_rule.get("idPoolType")
                if api_url:
                    processed_data, max_id_api = fetch_and_process_api_data_for_entity(api_url, entity_name, entity_rule, app_config_settings)
                    api_data_for_comparison[entity_name] = processed_data
                    if id_pool == 'dn': overall_max_dn_id_recomp = max(overall_max_dn_id_recomp, max_id_api)
                    elif id_pool == 'agent_group': overall_max_ag_id_recomp = max(overall_max_ag_id_recomp, max_id_api)
        logger.info(f"Re-compare Max IDs: DN={overall_max_dn_id_recomp}, AG={overall_max_ag_id_recomp}")

        output_workbook = openpyxl.load_workbook(processed_filepath, read_only=False, data_only=False)
        for entity_name_to_clear in api_data_for_comparison.keys():
            comp_sheet_to_clear = f"{entity_name_to_clear} Comparison"
            if comp_sheet_to_clear in output_workbook.sheetnames: del output_workbook[comp_sheet_to_clear]
        if METADATA_SHEET_NAME in output_workbook.sheetnames: del output_workbook[METADATA_SHEET_NAME]

        write_comparison_sheets(output_workbook, sheet_data_for_comparison_recomp, api_data_for_comparison, intermediate_data_recomp)

        metadata_sheet = output_workbook.create_sheet(title=METADATA_SHEET_NAME)
        metadata_sheet[MAX_DN_ID_LABEL_CELL] = "Max DN API ID (Comparison Run)"; metadata_sheet[MAX_DN_ID_LABEL_CELL].font = Font(bold=True)
        metadata_sheet[MAX_DN_ID_VALUE_CELL] = overall_max_dn_id_recomp
        metadata_sheet[MAX_AG_ID_LABEL_CELL] = "Max AgentGroup API ID (Comparison Run)"; metadata_sheet[MAX_AG_ID_LABEL_CELL].font = Font(bold=True)
        metadata_sheet[MAX_AG_ID_VALUE_CELL] = overall_max_ag_id_recomp
        output_workbook.save(processed_filepath)
        logger.info(f"Updated '{processed_filepath}' with new comparison and metadata.")

        if read_comparison_data(processed_filepath):
            logger.info(f"Re-loaded data from '{processed_filename}' into cache after re-comparison.")
            first_sheet_to_view = current_app.config.get('COMPARISON_SHEETS', [None])[0]
            if not first_sheet_to_view and current_app.config.get('EXCEL_DATA'): first_sheet_to_view = list(current_app.config['EXCEL_DATA'].keys())[0]
            return jsonify({ "message": f"Successfully re-compared data from '{processed_filename}'.", "redirect_url": url_for('ui.view_comparison', comparison_type=first_sheet_to_view) if first_sheet_to_view else url_for('ui.upload_config_page') }), 200
        else:
            return jsonify({"error": f"Comparison complete for '{processed_filename}', but failed to reload its data. Check logs."}), 500

    except Exception as e:
        logger.error(f"Error loading/comparing processed file '{processed_filename}': {e}", exc_info=True)
        return jsonify({"error": f"Error loading/comparing processed file: {e}"}), 500
    finally:
        if output_workbook:
            try: output_workbook.close()
            except Exception as e_close: logger.warning(f"Error closing workbook during load-processed: {e_close}")


@processing_bp.route('/update-config', methods=['POST'])
def update_config():
    """ API endpoint to save updated configuration data to config.ini. """
    logger.info("Received request to update configuration.")
    try:
        settings_to_save = {
            'api_timeout': request.form.get('timeout', type=int, default=15),
            'ideal_agent_header_text': request.form.get('ideal_agent_header_text'),
            'ideal_agent_fallback_cell': request.form.get('ideal_agent_fallback_cell'),
            'vag_extraction_sheet': request.form.get('vag_extraction_sheet'),
            'log_level_str': request.form.get('log_level')
        }
        if settings_to_save.get('api_timeout') is None: settings_to_save['api_timeout'] = 15
        if settings_to_save.get('log_level_str') is None: settings_to_save['log_level_str'] = 'INFO'

        config_path = current_app.config.get('CONFIG_FILE_PATH', 'config.ini')
        save_config(config_path, settings_to_save)
        current_app.config['APP_SETTINGS'].update(settings_to_save)
        logger.info("Configuration saved and application cache updated.")
        flash('Configuration saved successfully to config.ini. Restart may be needed for some changes.', 'success')
    except (IOError, ValueError, Exception) as e:
        logger.error(f"Error saving configuration: {e}", exc_info=True)
        flash(f'Error saving configuration: {e}', 'error')
    return redirect(url_for('ui.upload_config_page'))


@processing_bp.route('/simulate-configuration', methods=['POST'])
def simulate_configuration():
    """ API endpoint to simulate applying a DB update template to selected rows. """
    logger.info("Request received for /api/simulate-configuration")
    try:
        request_data = request.get_json();
        if not request_data: logger.warning("Simulate config: Invalid/empty JSON."); return jsonify({"error": "Invalid JSON payload."}), 400
        template_name = request_data.get('templateName'); selected_row_identifiers = request_data.get('selectedRowsData')
        if not template_name or selected_row_identifiers is None: logger.warning("Simulate config: Missing params."); return jsonify({"error": "Missing params."}), 400
        if not isinstance(selected_row_identifiers, list): logger.warning("'selectedRowsData' not a list."); return jsonify({"error": "'selectedRowsData' must be a list."}), 400
        logger.info(f"Simulating template '{template_name}' for {len(selected_row_identifiers)} items.")
        template_path = os.path.join(TEMPLATE_DIR, template_name)
        if not os.path.exists(template_path): logger.error(f"Template not found: {template_path}"); return jsonify({"error": f"Template '{template_name}' not found."}), 404
        try:
            with open(template_path, 'r', encoding='utf-8') as f: template_json = json.load(f)
        except Exception as e: logger.error(f"Error reading/parsing template {template_name}: {e}", exc_info=True); return jsonify({"error": f"Could not load/parse template '{template_name}'."}), 500
        
        all_excel_data = current_app.config.get('EXCEL_DATA', {}); sheet_headers_map = current_app.config.get('SHEET_HEADERS', {}); rows_to_process = []
        processed_identifiers = set()
        for sheet_name, sheet_data in all_excel_data.items():
            headers = sheet_headers_map.get(sheet_name)
            if not headers: continue
            id_key = headers[0]
            
            entity_type_for_id_gen = None
            if "vq" in sheet_name.lower(): entity_type_for_id_gen = 'dn'
            elif any(s_type in sheet_name.lower() for s_type in ["skill", "vag", "expr"]): entity_type_for_id_gen = 'agent_group'

            for row in sheet_data:
                row_identifier = row.get(id_key)
                if row_identifier in selected_row_identifiers and row_identifier not in processed_identifiers: rows_to_process.append((row, entity_type_for_id_gen)); processed_identifiers.add(row_identifier)
        
        found_count = len(rows_to_process); missing_identifiers = set(selected_row_identifiers) - processed_identifiers; missing_count = len(missing_identifiers)
        logger.info(f"Retrieved data for {found_count} of {len(selected_row_identifiers)} identifiers.")
        if missing_count > 0: logger.warning(f"Could not find data for identifiers: {missing_identifiers}")
        
        generated_payloads = []; processing_errors = []
        id_generator = IdGenerator(max_dn_id=current_app.config.get('MAX_DN_ID', 0), max_ag_id=current_app.config.get('MAX_AG_ID', 0))
        
        for row_data, entity_type_for_id in rows_to_process:
            first_header = list(row_data.keys())[0] if row_data else 'UNKNOWN'; row_id_for_log = row_data.get(first_header, "UNKNOWN_ID")
            try:
                current_row_id = None
                if entity_type_for_id == 'dn': current_row_id = id_generator.get_next_dn_id()
                elif entity_type_for_id == 'agent_group': current_row_id = id_generator.get_next_ag_id()
                else: logger.warning(f"Cannot generate ID for row '{row_id_for_log}' - unknown entity type '{entity_type_for_id}'.")
                generated_payload = replace_placeholders(template_json, row_data, current_row_id)
                generated_payloads.append(generated_payload)
            except Exception as e: logger.error(f"Error processing template for row '{row_id_for_log}': {e}", exc_info=True); processing_errors.append(f"Row '{row_id_for_log}': {e}")
        
        response_status_code = 200
        response_data = { "message": f"Simulation complete. Generated {len(generated_payloads)} payloads.", "status": "Simulation Success", "processed_count": found_count, "payloads": generated_payloads, "errors": [] }
        if missing_count > 0: response_data["message"] += f" Could not find data for {missing_count} identifiers: {list(missing_identifiers)}."; response_data["status"] = "Simulation Partial Success / Missing Data"; response_status_code = 207
        if processing_errors: response_data["errors"] = [str(e) for e in processing_errors]; response_data["message"] += f" Encountered {len(processing_errors)} errors."; response_data["status"] = "Simulation Partial Success / Errors" if response_status_code == 200 else response_data["status"]; response_status_code = 207
        logger.info(f"Simulation successful. Returning {len(generated_payloads)} payloads.")
        return jsonify(response_data), response_status_code
    except Exception as e: logger.error(f"Unexpected error in /api/simulate-configuration: {e}", exc_info=True); return jsonify({"error": "Internal server error during simulation."}), 500

@processing_bp.route('/confirm-update', methods=['POST'])
def confirm_update():
    """ API endpoint to receive previously generated payloads and perform (simulated) DB update. """
    logger.info("Request received for /api/confirm-update")
    try:
        request_data = request.get_json();
        if not request_data: logger.warning("Confirm update: Invalid/empty JSON."); return jsonify({"error": "Invalid JSON payload."}), 400
        payloads_to_commit = request_data.get('payloads')
        if payloads_to_commit is None or not isinstance(payloads_to_commit, list): logger.warning("Confirm update: Missing 'payloads' list."); return jsonify({"error": "Missing 'payloads' list or invalid format."}), 400
        logger.info(f"Received {len(payloads_to_commit)} payloads for final (simulated) update.")
        commit_errors = []; commit_success_count = 0
        logger.info(f"--- SIMULATING FINAL DATABASE UPDATE (START) ---")
        for i, payload in enumerate(payloads_to_commit):
            try: logger.info(f"Simulating DB Update for Payload {i+1}: {json.dumps(payload, indent=2)}"); commit_success_count += 1
            except Exception as db_err: logger.error(f"Simulated DB update FAILED for Payload {i+1}: {db_err}", exc_info=True); commit_errors.append(f"Payload {i+1}: {db_err}")
        logger.info(f"--- SIMULATING FINAL DATABASE UPDATE (END) ---")
        response_status_code = 200
        response_data = { "message": f"Simulated update completed for {commit_success_count} of {len(payloads_to_commit)} payloads.", "status": "Update Simulation Success", "success_count": commit_success_count, "error_count": len(commit_errors), "errors": [str(e) for e in commit_errors] }
        if commit_errors: response_data["status"] = "Update Simulation Partial Success / Errors"; response_status_code = 207
        if commit_success_count == 0 and len(payloads_to_commit) > 0: response_data["status"] = "Update Simulation Failed"; response_status_code = 500
        return jsonify(response_data), response_status_code

    except Exception as e: logger.error(f"Unexpected error in /api/confirm-update: {e}", exc_info=True); return jsonify({"error": "An internal server error occurred during update confirmation."}), 500

