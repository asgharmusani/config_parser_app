# -*- coding: utf-8 -*-
"""
Core Rule Engine for parsing Excel files based on user-defined rule templates.

This module defines the ExcelRuleEngine class which takes a set of rules
and an Excel workbook, then extracts data according to those rules.
The processing iterates through sheets, then cells once, applying the first
matching rule to a cell for primary entity identification.
Identifier rules are pre-processed for efficiency.
Uses a shared 'match_identifier_logic' from utils.py.
Stores cell coordinates instead of full cell objects for style reference where possible.
"""

import logging
import re
import openpyxl
from openpyxl.utils import cell as openpyxl_cell_utils
from typing import Dict, Any, List, Optional, Union, Set, Tuple

# --- MODIFICATION START: Import shared identifier matching logic ---
try:
    from utils import match_identifier_logic
except ImportError:
    logger = logging.getLogger(__name__) # Define logger early for this fallback
    logger.error("Failed to import match_identifier_logic from utils.py in excel_rule_engine.py")
    # Define a dummy function if utils.py or the function is missing
    def match_identifier_logic(value_to_check_str: str, identifier_rule: Dict[str, Any]) -> bool:
        """Dummy identifier matching function if import fails."""
        logger.error("Dummy match_identifier_logic called. Real function not imported from utils.py.")
        return False
# --- MODIFICATION END ---

logger = logging.getLogger(__name__) # Use module-specific logger

class ExcelRuleEngine:
    """
    Parses an Excel workbook based on a provided rule template (JSON).
    """

    def __init__(self, rule_template: Dict[str, Any]):
        """
        Initializes the ExcelRuleEngine with a set of rules.

        Args:
            rule_template: A dictionary representing the parsed JSON rule template.
                           Expected structure: {"Entities": [rule1, rule2, ...], "GlobalSettings": {...}}

        Raises:
            ValueError: If the rule template is malformed or missing essential parts.
        """
        logger.info("Initializing ExcelRuleEngine...")
        if not isinstance(rule_template, dict) or "Entities" not in rule_template:
            msg = "Rule template must be a dictionary with an 'Entities' key."
            logger.error(msg)
            raise ValueError(msg)

        self.rules: List[Dict[str, Any]] = rule_template["Entities"]
        if not isinstance(self.rules, list):
            msg = "The 'Entities' key in the rule template must contain a list of rules."
            logger.error(msg)
            raise ValueError(msg)

        # Load global settings with defaults
        self.global_settings = rule_template.get("GlobalSettings", {})
        self.default_skip_sheets = set(self.global_settings.get("defaultSkipSheets", ["Metadata", "Instructions", "Summary"]))
        self.global_default_check_for_strikethrough = self.global_settings.get("defaultCheckForStrikethrough", False)

        # Validate and pre-process rules for efficiency
        self._validate_and_preprocess_rules()
        logger.info(f"ExcelRuleEngine initialized with {len(self.rules)} entity rules.")

    def _validate_and_preprocess_rules(self):
        """
        Performs basic validation on the loaded entity rules and pre-processes
        identifier components for efficiency.
        Checks for required keys in each rule and validates their types.
        """
        logger.debug("Validating and pre-processing entity rules...")
        for i, rule in enumerate(self.rules):
            if not isinstance(rule, dict):
                raise ValueError(f"Rule at index {i} is not a dictionary.")
            rule_name_for_error = rule.get('name', f'Unnamed rule at index {i}')

            # --- Mandatory keys ---
            required_keys = ["name", "identifier"] # 'enabled' defaults to True if not present
            for key in required_keys:
                if key not in rule:
                    raise ValueError(f"Rule '{rule_name_for_error}' is missing required key: '{key}'.")
            if not isinstance(rule["name"], str) or not rule["name"]:
                 raise ValueError(f"Rule name for rule at index {i} must be a non-empty string.")

            # --- Identifier validation and pre-processing ---
            identifier = rule["identifier"]
            if not isinstance(identifier, dict):
                raise ValueError(f"Identifier for rule '{rule_name_for_error}' must be a dictionary.")

            id_type_raw = identifier.get("type")
            id_value = identifier.get("value")
            valid_id_types = ["startswith", "contains", "exactmatch", "regex"]

            if not id_type_raw or not isinstance(id_type_raw, str):
                 raise ValueError(f"Identifier 'type' for rule '{rule_name_for_error}' must be a non-empty string.")
            id_type_lower = id_type_raw.lower()
            if id_type_lower not in valid_id_types:
                raise ValueError(f"Invalid identifier type '{id_type_raw}' for rule '{rule_name_for_error}'. Must be one of {valid_id_types}.")
            if not isinstance(id_value, str) or not id_value: # Value must be non-empty string
                raise ValueError(f"Identifier 'value' for rule '{rule_name_for_error}' must be a non-empty string.")

            # Store pre-processed identifier components directly in the rule's identifier dict
            # This is used by the match_identifier_logic in utils.py
            identifier['_type_processed'] = id_type_lower
            identifier['_value_original'] = id_value # Keep original for regex
            case_sensitive = identifier.get("caseSensitive", False) # Default false for non-regex
            if not isinstance(case_sensitive, bool):
                 raise ValueError(f"Optional 'identifier.caseSensitive' for rule '{rule_name_for_error}' must be a boolean.")
            identifier['_case_sensitive_processed'] = case_sensitive

            if id_type_lower != "regex" and not case_sensitive:
                identifier['_value_to_compare_processed'] = id_value.lower()
            else:
                identifier['_value_to_compare_processed'] = id_value

            if id_type_lower == "regex":
                try:
                    identifier['_compiled_regex_processed'] = re.compile(id_value)
                except re.error as e:
                    raise ValueError(f"Invalid regex pattern '{id_value}' for rule '{rule_name_for_error}': {e}")
            # --- End Identifier Pre-processing ---

            # --- Optional keys validation (no changes from previous version) ---
            # ... (keep existing validation for sheets, primaryFieldKey, comparisonApiUrl, idPoolType, etc.) ...
            if "sheets" in rule and not (rule["sheets"] is None or (isinstance(rule["sheets"], list) and all(isinstance(s, str) for s in rule["sheets"]))):
                raise ValueError(f"Optional 'sheets' key for rule '{rule_name_for_error}' must be null or a list of strings.")
            if "primaryFieldKey" in rule and (not isinstance(rule["primaryFieldKey"], str) or not rule["primaryFieldKey"]):
                raise ValueError(f"Optional 'primaryFieldKey' for rule '{rule_name_for_error}' must be a non-empty string.")
            if "comparisonApiUrl" in rule and (rule["comparisonApiUrl"] is not None and not isinstance(rule["comparisonApiUrl"], str)):
                raise ValueError(f"Optional 'comparisonApiUrl' for rule '{rule_name_for_error}' must be a string or null.")
            if "idPoolType" in rule and rule["idPoolType"] not in [None, "dn", "agent_group"]: # Allow null
                raise ValueError(f"Optional 'idPoolType' for rule '{rule_name_for_error}' must be 'dn', 'agent_group', or null.")
            if "replaceRules" in rule:
                if not isinstance(rule["replaceRules"], list): raise ValueError(f"'replaceRules' for rule '{rule_name_for_error}' must be a list.")
                for rep_rule_idx, rep_rule in enumerate(rule["replaceRules"]):
                    if not (isinstance(rep_rule, dict) and "find" in rep_rule and "replace" in rep_rule and isinstance(rep_rule["find"], str) and isinstance(rep_rule["replace"], str)):
                        raise ValueError(f"Invalid replaceRule at index {rep_rule_idx} for rule '{rule_name_for_error}'.")
            if "fetchAdditionalColumn" in rule:
                add_col_rule = rule["fetchAdditionalColumn"]
                if not isinstance(add_col_rule, dict): raise ValueError(f"'fetchAdditionalColumn' for rule '{rule_name_for_error}' must be a dictionary.")
                required_add_col_keys = ["targetKeyName", "searchHeaderName", "searchInLocations"]
                for key_ac in required_add_col_keys:
                    if key_ac not in add_col_rule: raise ValueError(f"'fetchAdditionalColumn' for rule '{rule_name_for_error}' missing key: '{key_ac}'.")
                if not (isinstance(add_col_rule["searchInLocations"], list) and all(isinstance(s, str) for s in add_col_rule["searchInLocations"])):
                     raise ValueError(f"'searchInLocations' for 'fetchAdditionalColumn' in rule '{rule_name_for_error}' must be a list of strings.")
                if "replaceRules" in add_col_rule and not isinstance(add_col_rule["replaceRules"], list):
                        raise ValueError(f"'replaceRules' in 'fetchAdditionalColumn' for rule '{rule_name_for_error}' must be a list.")
                if "valueFromRowOffset" in add_col_rule and not isinstance(add_col_rule["valueFromRowOffset"], int):
                    raise ValueError(f"Optional 'valueFromRowOffset' in 'fetchAdditionalColumn' for rule '{rule_name_for_error}' must be an integer.")
            if "sourceFromField" in rule:
                if not isinstance(rule["sourceFromField"], str) or '.' not in rule["sourceFromField"]:
                    raise ValueError(f"'sourceFromField' for rule '{rule_name_for_error}' must be a string in 'ParentEntityName.FieldName' format.")
                if identifier.get("_type_processed") != "regex":
                    logger.warning(f"Rule '{rule_name_for_error}' uses 'sourceFromField'. Its 'identifier' type is '{identifier.get('type')}'. Consider 'regex' type identifier with capturing group for multiple extractions.")
            if "extractSubEntities" in rule:
                sub_entity_rule = rule["extractSubEntities"]
                if not isinstance(sub_entity_rule, dict): raise ValueError(f"'extractSubEntities' for rule '{rule_name_for_error}' must be a dictionary.")
                required_sub_keys = ["subEntityName", "regex"]
                for key_se in required_sub_keys:
                    if key_se not in sub_entity_rule: raise ValueError(f"'extractSubEntities' for rule '{rule_name_for_error}' missing key: '{key_se}'.")
                if "sourceValueFrom" in sub_entity_rule and sub_entity_rule["sourceValueFrom"] not in ["primaryFieldKey"] and not sub_entity_rule["sourceValueFrom"].startswith("additional."):
                    raise ValueError(f"Invalid 'sourceValueFrom' in 'extractSubEntities' for rule '{rule_name_for_error}'.")
                if "replaceRules" in sub_entity_rule and not isinstance(sub_entity_rule["replaceRules"], list):
                    raise ValueError(f"'replaceRules' in 'extractSubEntities' for rule '{rule_name_for_error}' must be a list.")
            if "constructFields" in rule:
                if not isinstance(rule["constructFields"], list): raise ValueError(f"'constructFields' for rule '{rule_name_for_error}' must be a list.")
                for con_field_idx, con_field_rule in enumerate(rule["constructFields"]):
                    if not isinstance(con_field_rule, dict): raise ValueError(f"Item at index {con_field_idx} in 'constructFields' for rule '{rule_name_for_error}' must be a dictionary.")
                    required_con_keys = ["targetKeyName", "formatString"]
                    for key_cf in required_con_keys:
                        if key_cf not in con_field_rule: raise ValueError(f"Item at index {con_field_idx} in 'constructFields' for rule '{rule_name_for_error}' missing key: '{key_cf}'.")
                    if "onMissingSource" in con_field_rule and con_field_rule["onMissingSource"] not in ["skip_field", "empty_string", "error"]:
                        raise ValueError(f"Invalid 'onMissingSource' for 'constructFields' item in rule '{rule_name_for_error}'.")
        logger.debug("All rules passed basic validation and identifier pre-processing.")


    def _apply_replace_rules(self, text_value: str, replace_rules: List[Dict[str, str]]) -> str:
        """Applies a list of find/replace rules to a given text value."""
        if not isinstance(text_value, str):
            return text_value
        processed_value = text_value
        for rule_item in replace_rules:
            find_str = rule_item.get("find", "")
            replace_str = rule_item.get("replace", "")
            processed_value = processed_value.replace(find_str, replace_str)
        return processed_value

    # --- MODIFICATION START: _match_identifier method is REMOVED ---
    # It will now use the match_identifier_logic from utils.py
    # --- MODIFICATION END ---

    def _find_additional_column_header_once_per_sheet(
            self,
            sheet: openpyxl.worksheet.worksheet.Worksheet,
            fetch_additional_column_rule: Dict[str, Any],
            sheet_header_cache: Dict[str, Optional[int]]
        ) -> Optional[int]:
        """
        Finds and caches the column index for an additional column's header.
        (No changes from previous version of this method)
        """
        search_header_name = fetch_additional_column_rule.get("searchHeaderName")
        search_in_locations = fetch_additional_column_rule.get("searchInLocations", [])
        if not search_header_name: logger.warning("fetchAdditionalColumn rule missing 'searchHeaderName'."); return None
        cache_key = search_header_name
        if cache_key in sheet_header_cache: return sheet_header_cache[cache_key]
        found_column_idx = None
        for loc in search_in_locations:
            try:
                if re.fullmatch(r'[A-Z]+', loc, re.IGNORECASE):
                    col_idx_from_letter = openpyxl_cell_utils.column_index_from_string(loc)
                    if col_idx_from_letter <= sheet.max_column:
                        header_cell_value = sheet.cell(row=1, column=col_idx_from_letter).value
                        if header_cell_value and search_header_name in str(header_cell_value): found_column_idx = col_idx_from_letter; break
                elif re.fullmatch(r'[A-Z]+[1-9][0-9]*', loc, re.IGNORECASE):
                    col_str, row_str = openpyxl_cell_utils.coordinate_to_tuple(loc)
                    header_row_idx, header_col_idx = int(row_str), openpyxl_cell_utils.column_index_from_string(col_str)
                    if header_row_idx <= sheet.max_row and header_col_idx <= sheet.max_column:
                        header_cell_value = sheet.cell(row=header_row_idx, column=header_col_idx).value
                        if header_cell_value and search_header_name in str(header_cell_value): found_column_idx = header_col_idx; break
            except Exception as e: logger.warning(f"Error processing searchIn location '{loc}' for '{search_header_name}' on sheet '{sheet.title}': {e}")
        sheet_header_cache[cache_key] = found_column_idx
        log_msg = f"Header '{search_header_name}' found in column {found_column_idx}" if found_column_idx else f"Header '{search_header_name}' not found"
        logger.debug(f"{log_msg} for sheet '{sheet.title}'. Caching result.")
        return found_column_idx

    def _fetch_additional_column_data_from_row(
        self, current_row_idx: int, sheet: openpyxl.worksheet.worksheet.Worksheet,
        found_column_idx: int, replace_rules: List[Dict[str, str]], value_from_row_offset: int = 0
    ) -> Optional[str]:
        """ Fetches and cleans data from a specific cell, potentially offset from the current row. """
        # (No changes from previous version of this method)
        target_row_idx = current_row_idx + value_from_row_offset
        if not (1 <= target_row_idx <= sheet.max_row): logger.warning(f"Target row {target_row_idx} out of bounds for sheet '{sheet.title}'."); return None
        if found_column_idx > sheet.max_column: logger.warning(f"Target column {found_column_idx} exceeds max column {sheet.max_column} for sheet '{sheet.title}'."); return None
        additional_cell_value = sheet.cell(row=target_row_idx, column=found_column_idx).value
        if additional_cell_value is not None:
            value_str = str(additional_cell_value).strip()
            if replace_rules: value_str = self._apply_replace_rules(value_str, replace_rules)
            return value_str
        return None

    def _extract_sub_entities(
        self, source_text: str, sub_entity_rule: Dict[str, Any], primary_cell_strikethrough: bool
    ) -> List[Dict[str, Any]]:
        """ Extracts sub-entities from a source text using regex. """
        # (No changes from previous version of this method)
        if not source_text or not isinstance(source_text, str): return []
        regex_pattern = sub_entity_rule.get("regex")
        replace_rules = sub_entity_rule.get("replaceRules", [])
        apply_primary_strike = sub_entity_rule.get("checkForStrikethrough", False)
        final_strike_status = primary_cell_strikethrough if apply_primary_strike else False
        if not regex_pattern: logger.warning("Missing 'regex' in 'extractSubEntities' rule."); return []
        extracted_values = []
        try:
            compiled_re = re.compile(regex_pattern)
            matches = compiled_re.findall(source_text)
            for match_value in matches:
                value_to_clean = match_value[0] if isinstance(match_value, tuple) and match_value else match_value
                cleaned_value = str(value_to_clean).strip()
                if replace_rules: cleaned_value = self._apply_replace_rules(cleaned_value, replace_rules)
                if cleaned_value: extracted_values.append({"value": cleaned_value, "strike": final_strike_status})
        except re.error as e: logger.error(f"Invalid regex '{regex_pattern}' in 'extractSubEntities': {e}")
        except Exception as e: logger.error(f"Error extracting sub-entities with regex '{regex_pattern}': {e}", exc_info=True)
        return extracted_values

    def _construct_field(
            self, format_string: str, current_entity_data: Dict[str, Any],
            primary_key_name_for_this_rule: str, on_missing_source: str
        ) -> Optional[str]:
        """ Constructs a new field value based on a format string and existing entity data. """
        # (No changes from previous version of this method)
        placeholder_pattern = re.compile(r'{([^}]+)}')
        def replace_match(match):
            field_name_to_lookup = match.group(1).strip()
            if field_name_to_lookup == "_primary_": return str(current_entity_data.get(primary_key_name_for_this_rule, ""))
            elif field_name_to_lookup in current_entity_data: return str(current_entity_data.get(field_name_to_lookup, ""))
            else:
                if on_missing_source == "empty_string": logger.debug(f"Missing source field '{field_name_to_lookup}' for constructField, using empty string."); return ""
                elif on_missing_source == "error": logger.error(f"Missing source field '{field_name_to_lookup}' for constructField (onMissingSource=error)."); raise KeyError(f"Missing source field '{field_name_to_lookup}'")
                elif on_missing_source == "skip_field": logger.debug(f"Missing source field '{field_name_to_lookup}' for constructField, will skip field."); raise ValueError("_SKIP_CONSTRUCT_FIELD_")
                return match.group(0)
        try: return placeholder_pattern.sub(replace_match, format_string)
        except ValueError as e:
            if str(e) == "_SKIP_CONSTRUCT_FIELD_": return None
            raise
        except KeyError as e: raise


    def process_workbook(self, workbook: openpyxl.workbook.Workbook) -> Dict[str, List[Dict[str, Any]]]:
        """
        Processes the entire workbook based on the loaded rules.
        Iterates sheets first, then cells once per sheet. For each cell, it tries to find
        a matching rule. A cell is "claimed" as a primary entity by the first rule that identifies it.

        Args:
            workbook: An openpyxl Workbook object.

        Returns:
            A dictionary where keys are entity names (from rules) and values are lists
            of extracted row data dictionaries.
        """
        logger.info(f"Processing workbook with {len(self.rules)} rules using CELL-FIRST, RULE-DISPATCH optimized approach.")
        parsed_entities: Dict[str, List[Dict[str, Any]]] = {}
        sheet_header_location_cache: Dict[str, Dict[str, Optional[int]]] = {}
        claimed_primary_cells: Set[Tuple[str, int, int]] = set()

        for rule in self.rules:
            if rule.get("enabled", True):
                parsed_entities[rule["name"]] = []

        # --- PASS 1: Process rules that identify entities directly from Excel cells ---
        logger.info("Rule Engine - PASS 1: Processing direct Excel cell identifiers...")
        for sheet in workbook.worksheets:
            is_globally_skipped = sheet.title in self.default_skip_sheets
            if is_globally_skipped:
                is_explicitly_included = any(rule_check.get("enabled", True) and sheet.title in rule_check.get("sheets", []) for rule_check in self.rules if rule_check.get("sheets") is not None)
                if not is_explicitly_included:
                    logger.info(f"Skipping sheet (globally): {sheet.title}")
                    continue
            logger.info(f"PASS 1 - Processing sheet: {sheet.title} (Max Row: {sheet.max_row}, Max Col: {sheet.max_column})")
            if sheet.title not in sheet_header_location_cache:
                sheet_header_location_cache[sheet.title] = {}

            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell_coordinate_tuple = (sheet.title, row_idx, col_idx)
                    if cell_coordinate_tuple in claimed_primary_cells: continue
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is None or str(cell.value).strip() == "": continue
                    cell_value_str = str(cell.value).strip()

                    for rule in self.rules:
                        if not rule.get("enabled", True) or "sourceFromField" in rule: continue
                        rule_sheets_filter = rule.get("sheets")
                        if rule_sheets_filter is not None and sheet.title not in rule_sheets_filter: continue

                        # --- MODIFICATION: Use imported match_identifier_logic ---
                        if match_identifier_logic(cell_value_str, rule["identifier"]):
                        # --- END MODIFICATION ---
                            logger.debug(f"PASS 1 MATCH: Rule '{rule['name']}', Cell {cell.coordinate}")
                            claimed_primary_cells.add(cell_coordinate_tuple)
                            primary_value = cell_value_str
                            rule_check_strike = rule["identifier"].get("checkForStrikethrough", self.global_default_check_for_strikethrough)
                            primary_strike_status = bool(cell.font and cell.font.strike if rule_check_strike else False)
                            if "replaceRules" in rule: primary_value = self._apply_replace_rules(primary_value, rule["replaceRules"])
                            entity_data: Dict[str, Any] = {}
                            primary_key_name = rule.get("primaryFieldKey", rule["name"])
                            entity_data[primary_key_name] = primary_value
                            entity_data["strike"] = primary_strike_status
                            entity_data["_source_sheet_title_"] = sheet.title
                            entity_data["_source_cell_coordinate_"] = cell.coordinate
                            entity_data["_rule_primary_field_key_"] = primary_key_name

                            if "fetchAdditionalColumn" in rule:
                                add_col_config = rule["fetchAdditionalColumn"]
                                header_col_idx_found = self._find_additional_column_header_once_per_sheet(sheet, add_col_config, sheet_header_location_cache[sheet.title])
                                if header_col_idx_found:
                                    offset = add_col_config.get("valueFromRowOffset", 0)
                                    additional_value = self._fetch_additional_column_data_from_row(row_idx, sheet, header_col_idx_found, add_col_config.get("replaceRules", []), offset)
                                    if additional_value is not None:
                                        target_key = add_col_config.get("targetKeyName")
                                        if target_key: entity_data[target_key] = additional_value
                                        else: logger.warning(f"Missing 'targetKeyName' in fetchAdditionalColumn for rule '{rule['name']}'.")
                            if "constructFields" in rule:
                                for construct_rule_config in rule["constructFields"]:
                                    target_key = construct_rule_config.get("targetKeyName"); format_string = construct_rule_config.get("formatString"); on_missing = construct_rule_config.get("onMissingSource", "skip_field")
                                    if target_key and format_string:
                                        try:
                                            constructed_value = self._construct_field(format_string, entity_data, primary_key_name, on_missing)
                                            if constructed_value is not None: entity_data[target_key] = constructed_value
                                        except KeyError as e: logger.error(f"Error constructing field '{target_key}' for rule '{rule['name']}': {e}")
                                    else: logger.warning(f"Missing 'targetKeyName' or 'formatString' in constructFields for rule '{rule['name']}'.")
                            if "extractSubEntities" in rule :
                                sub_entity_config = rule["extractSubEntities"]
                                source_value_for_sub_extraction = ""
                                source_value_key_from_rule = sub_entity_config.get("sourceValueFrom", "primaryFieldKey")
                                if source_value_key_from_rule == "primaryFieldKey": source_value_for_sub_extraction = entity_data.get(primary_key_name, "")
                                elif source_value_key_from_rule.startswith("additional."): additional_key = source_value_key_from_rule.split("additional.", 1)[1]; source_value_for_sub_extraction = entity_data.get(additional_key, "")
                                else: logger.warning(f"Invalid 'sourceValueFrom' ('{source_value_key_from_rule}') for sub-entity extraction in rule '{rule['name']}'."); source_value_for_sub_extraction = entity_data.get(primary_key_name, "")
                                if source_value_for_sub_extraction:
                                    sub_entities = self._extract_sub_entities(source_value_for_sub_extraction, sub_entity_config, primary_strike_status)
                                    if sub_entities:
                                        sub_entity_list_key = sub_entity_config.get("subEntityName")
                                        if sub_entity_list_key: entity_data[sub_entity_list_key] = sub_entities
                                        else: logger.warning(f"Missing 'subEntityName' in extractSubEntities rule for '{rule['name']}'.")
                            parsed_entities[rule["name"]].append(entity_data)
                            break # Cell claimed

        # --- PASS 2: Process rules that source data from other entities ---
        logger.info("Rule Engine - PASS 2: Processing rules with 'sourceFromField'...")
        for rule in self.rules:
            if not rule.get("enabled", True) or "sourceFromField" not in rule: continue
            logger.info(f"PASS 2 - Applying rule: '{rule['name']}' (sourced)")
            source_from_field_path = rule["sourceFromField"]
            try: parent_entity_name, source_field_key = source_from_field_path.split('.', 1)
            except ValueError: logger.error(f"Invalid 'sourceFromField' format '{source_from_field_path}' in rule '{rule['name']}'. Skipping."); continue
            if parent_entity_name not in parsed_entities: logger.warning(f"Parent entity '{parent_entity_name}' for rule '{rule['name']}' not found. Skipping."); continue

            for parent_entity_instance in parsed_entities[parent_entity_name]:
                source_string_from_parent = parent_entity_instance.get(source_field_key)
                parent_primary_strike_status = parent_entity_instance.get("strike", False)
                parent_source_sheet_title = parent_entity_instance.get("_source_sheet_title_")
                parent_source_cell_coord = parent_entity_instance.get("_source_cell_coordinate_")
                parent_primary_key_name = parent_entity_instance.get("_rule_primary_field_key_", parent_entity_name)

                if source_string_from_parent is None or not isinstance(source_string_from_parent, str):
                    logger.debug(f"Source field '{source_field_key}' in parent '{parent_entity_name}' is None or not string. Skipping."); continue

                identifier_rule = rule["identifier"]
                id_type = identifier_rule.get('_type_processed')
                compiled_regex_child = identifier_rule.get('_compiled_regex_processed')
                extracted_values_for_child = []

                if id_type == "regex" and compiled_regex_child:
                    matches = compiled_regex_child.findall(source_string_from_parent)
                    for match in matches:
                        extracted_value = match[0] if isinstance(match, tuple) and match else match
                        if isinstance(extracted_value, str): extracted_values_for_child.append(extracted_value.strip())
                # --- MODIFICATION: Use imported match_identifier_logic ---
                elif match_identifier_logic(source_string_from_parent, identifier_rule):
                # --- END MODIFICATION ---
                    extracted_values_for_child.append(source_string_from_parent)

                for primary_value_for_child in extracted_values_for_child:
                    if not primary_value_for_child: continue
                    rule_check_strike_child = identifier_rule.get("checkForStrikethrough", self.global_default_check_for_strikethrough)
                    child_strike_status = parent_primary_strike_status if rule_check_strike_child else False
                    if "replaceRules" in rule: primary_value_for_child = self._apply_replace_rules(primary_value_for_child, rule["replaceRules"])
                    child_entity_data: Dict[str, Any] = {}
                    child_primary_key_name = rule.get("primaryFieldKey", rule["name"])
                    child_entity_data[child_primary_key_name] = primary_value_for_child
                    child_entity_data["strike"] = child_strike_status
                    child_entity_data["_source_sheet_title_"] = parent_source_sheet_title
                    child_entity_data["_source_cell_coordinate_"] = parent_source_cell_coord
                    child_entity_data["_parent_entity_"] = parent_entity_name
                    child_entity_data["_parent_key_"] = parent_entity_instance.get(parent_primary_key_name)
                    child_entity_data["_rule_primary_field_key_"] = child_primary_key_name

                    if "fetchAdditionalColumn" in rule:
                        add_col_config = rule["fetchAdditionalColumn"]
                        if parent_source_sheet_title and parent_source_cell_coord:
                            parent_sheet_obj = workbook[parent_source_sheet_title]
                            parent_row_idx_for_add_col, _ = openpyxl_cell_utils.coordinate_to_tuple(parent_source_cell_coord)
                            header_col_idx_found = self._find_additional_column_header_once_per_sheet(parent_sheet_obj, add_col_config, sheet_header_location_cache[parent_sheet_title])
                            if header_col_idx_found:
                                offset = add_col_config.get("valueFromRowOffset", 0)
                                additional_value = self._fetch_additional_column_data_from_row(parent_row_idx_for_add_col, parent_sheet_obj, header_col_idx_found, add_col_config.get("replaceRules", []), offset)
                                if additional_value is not None:
                                    target_key = add_col_config.get("targetKeyName")
                                    if target_key: child_entity_data[target_key] = additional_value
                                    else: logger.warning(f"Missing 'targetKeyName' in fetchAdditionalColumn for sourced rule '{rule['name']}'.")
                    if "constructFields" in rule:
                        for construct_rule_config in rule["constructFields"]:
                            target_key = construct_rule_config.get("targetKeyName"); format_string = construct_rule_config.get("formatString"); on_missing = construct_rule_config.get("onMissingSource", "skip_field")
                            if target_key and format_string:
                                try:
                                    constructed_value = self._construct_field(format_string, child_entity_data, child_primary_key_name, on_missing)
                                    if constructed_value is not None: child_entity_data[target_key] = constructed_value
                                except KeyError as e: logger.error(f"Error constructing field '{target_key}' for sourced rule '{rule['name']}': {e}")
                            else: logger.warning(f"Missing 'targetKeyName' or 'formatString' in constructFields for sourced rule '{rule['name']}'.")
                    parsed_entities[rule["name"]].append(child_entity_data)
        logger.info("Finished processing workbook with rule engine.")
        return parsed_entities

