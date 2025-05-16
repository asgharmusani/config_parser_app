# -*- coding: utf-8 -*-
"""
Common utility functions used across the Excel Comparator application.
Includes Excel styling, placeholder replacement, ID generation,
reading processed comparison data from Excel, and identifier matching.
"""

import logging
import re
import os # For path manipulation
import openpyxl
from openpyxl.styles import Font, PatternFill # Ensure Font/PatternFill are imported if used
from openpyxl.utils import cell as openpyxl_cell_utils
from openpyxl.utils.exceptions import InvalidFileException # For specific exception handling
from typing import Optional, Any, Dict, Tuple, Set, List
from flask import current_app # For accessing app.config in read_comparison_data

logger = logging.getLogger(__name__) # Use module-specific logger

# --- Constants for read_comparison_data ---
# These should align with what excel_comparator.py (or the built-in parser) writes
COMPARISON_SUFFIX = " Comparison"
METADATA_SHEET_NAME = "Metadata"
MAX_DN_ID_VALUE_CELL = "B1" # Cell in Metadata sheet containing Max DN ID value
MAX_AG_ID_VALUE_CELL = "B2" # Cell in Metadata sheet containing Max AG ID value


# --- Excel Utilities ---
def copy_cell_style(source_cell: openpyxl.cell.Cell, target_cell: openpyxl.cell.Cell):
    """
    Copies font, fill, alignment, and number format style from source_cell to target_cell.

    Args:
        source_cell: The openpyxl Cell object to copy style from.
        target_cell: The openpyxl Cell object to copy style to.
    """
    if source_cell.has_style:
        # Copy Font properties
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        # Copy Fill properties
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        # Copy Alignment properties
        if source_cell.alignment:
            target_cell.alignment = openpyxl.styles.Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
        # Copy Number format
        target_cell.number_format = source_cell.number_format
    else:
        # Apply default styles if source has no specific style applied
        target_cell.font = Font()
        target_cell.fill = PatternFill()
        target_cell.alignment = openpyxl.styles.Alignment()
        target_cell.number_format = 'General'


def extract_skills(expression: str) -> list[str]:
    """
    Extracts potential skill names (alphanumeric + underscore) from a skill
    expression string. Looks for patterns like 'SkillName>5'.

    Args:
        expression: The skill expression string.

    Returns:
        A list of extracted skill names.
    """
    # Regex finds words (alphanumeric or underscore) followed immediately by '>' and one or more digits
    # The capturing group is the skill name itself.
    skills = re.findall(r'\b([a-zA-Z0-9_]+)(?=>\d+)', expression)
    # logger.debug(f"Extracted skills {skills} from expression '{expression}'") # Logging can be done by caller
    return skills


# --- ID Generation Helper ---
class IdGenerator:
    """
    Generates sequential IDs within a single request context, maintaining
    separate sequences for DN (VQ) and Agent Group entities based on
    Max IDs provided during initialization.
    """
    def __init__(self, max_dn_id: int = 0, max_ag_id: int = 0):
        """
        Initializes the generator with the starting max IDs.
        These max IDs should be fetched from app.config, which are read
        from the Metadata sheet of the currently loaded processed Excel file.

        Args:
            max_dn_id: The highest existing ID found for DN entities.
            max_ag_id: The highest existing ID found for Agent Group entities.
        """
        # The next ID should be one greater than the maximum found for each type
        self._next_dn_id = max_dn_id + 1
        self._next_ag_id = max_ag_id + 1
        logger.info(f"ID Generator initialized. Next DN ID: {self._next_dn_id} (based on max: {max_dn_id}), Next AG ID: {self._next_ag_id} (based on max: {max_ag_id})")

    def get_next_dn_id(self) -> int:
        """Returns the next sequential DN ID and increments the DN counter."""
        next_id = self._next_dn_id
        self._next_dn_id += 1
        logger.debug(f"Generated next DN ID: {next_id}")
        return next_id

    def get_next_ag_id(self) -> int:
        """Returns the next sequential Agent Group ID and increments the AG counter."""
        next_id = self._next_ag_id
        self._next_ag_id += 1
        logger.debug(f"Generated next AG ID: {next_id}")
        return next_id

# --- Template Placeholder Replacement ---
def replace_placeholders(template_data: Any, row_data: dict, current_row_next_id: Optional[int] = None) -> Any:
    """
    Recursively traverses a template structure (dict, list, or string)
    and replaces placeholders with values from row_data or the pre-generated ID.

    Supported Placeholders:
    - {row.ColumnName}: Replaced with the value from the corresponding column in row_data.
                        Lookup is case-insensitive for the ColumnName part.
    - {func.next_id}: Replaced with the current_row_next_id value.

    Args:
        template_data: The template structure (can be dict, list, string, etc.).
        row_data: The dictionary containing data for the current row (keys are actual headers).
        current_row_next_id: The pre-generated sequential ID for the current row.

    Returns:
        The template structure with placeholders replaced.
    """
    placeholder_pattern = re.compile(r'{(\w+)\.([^}]+)}')

    def perform_replace(text: str) -> str:
        """Performs replacements on a single string."""
        if not isinstance(text, str):
            return text

        def replace_match(match):
            placeholder_type = match.group(1).lower()
            placeholder_name = match.group(2).strip()

            if placeholder_type == 'row':
                found_key = None
                # Case-insensitive key lookup in row_data
                for key_in_row in row_data.keys():
                    if key_in_row.lower() == placeholder_name.lower():
                        found_key = key_in_row
                        break
                if found_key:
                    replacement = row_data.get(found_key, "")
                else:
                    replacement = ""
                    logger.warning(f"Placeholder {{row.{placeholder_name}}} not found in row data keys: {list(row_data.keys())}")
                return str(replacement)
            elif placeholder_type == 'func':
                if placeholder_name == 'next_id':
                    if current_row_next_id is not None:
                        return str(current_row_next_id)
                    else:
                        logger.warning(f"Placeholder {{func.next_id}} used but no ID provided for this row.")
                        return "{ERROR:next_id_missing}"
                else:
                    logger.warning(f"Unknown function placeholder: {match.group(0)}")
                    return match.group(0)
            else:
                 logger.warning(f"Unknown placeholder type in template: {match.group(0)}")
                 return match.group(0)
        return placeholder_pattern.sub(replace_match, text)

    if isinstance(template_data, str):
        return perform_replace(template_data)
    elif isinstance(template_data, dict):
        return {
            key: replace_placeholders(value, row_data, current_row_next_id)
            for key, value in template_data.items()
        }
    elif isinstance(template_data, list):
        return [
            replace_placeholders(item, row_data, current_row_next_id)
            for item in template_data
        ]
    else:
        return template_data


# --- Identifier Matching Logic (Shared) ---
def match_identifier_logic(value_to_check_str: str, identifier_rule: Dict[str, Any]) -> bool:
    """
    Checks if a string value matches a given identifier rule.
    This function is used by both ExcelRuleEngine and api_fetching.

    Args:
        value_to_check_str: The string value to check.
        identifier_rule: The identifier part of an entity rule.
                         It expects pre-processed keys if coming from ExcelRuleEngine's
                         `_validate_and_preprocess_rules` (e.g., '_type_processed'),
                         but will fall back to raw keys ('type', 'value', 'caseSensitive')
                         if pre-processed keys are not found (e.g., if called directly
                         with a raw rule snippet from api_fetching).

    Returns:
        True if the value matches the rule, False otherwise.
    """
    # Prefer pre-processed keys, but fallback to raw keys from rule definition
    id_type = identifier_rule.get('_type_processed', str(identifier_rule.get("type", "")).lower())
    case_sensitive = identifier_rule.get('_case_sensitive_processed', identifier_rule.get("caseSensitive", False))

    # Get the value to compare against from the rule
    if id_type == "regex":
        value_from_rule = identifier_rule.get('_value_original', identifier_rule.get("value", ""))
        compiled_regex = identifier_rule.get('_compiled_regex_processed')
        # If regex is not pre-compiled (e.g. direct call), compile it now
        if not compiled_regex and value_from_rule:
            try:
                compiled_regex = re.compile(value_from_rule)
            except re.error as e:
                logger.warning(f"Invalid regex '{value_from_rule}' in identifier rule during match: {e}")
                return False
    else: # For non-regex types
        value_from_rule = identifier_rule.get('_value_to_compare_processed', identifier_rule.get("value", ""))
        # If not pre-processed and not case-sensitive, make rule value lower for comparison
        if '_value_to_compare_processed' not in identifier_rule and not case_sensitive:
            value_from_rule = value_from_rule.lower()

    # Basic check for value existence
    if not value_from_rule and id_type != "regex":
        logger.debug(f"Identifier rule missing value for non-regex type: {identifier_rule}")
        return False
    if id_type == "regex" and not compiled_regex: # Regex must have a pattern
        logger.debug(f"Regex identifier rule missing value or compiled pattern: {identifier_rule}")
        return False

    # Prepare the value_to_check_str based on case sensitivity for non-regex types
    val_to_check_prepared = value_to_check_str
    if id_type != "regex" and not case_sensitive:
        val_to_check_prepared = value_to_check_str.lower()
    # For regex, value_from_rule is the pattern, and cell_value_str is used as is.
    # For case-sensitive non-regex, value_from_rule is used as is.

    # Perform the match
    if id_type == "startswith":
        return val_to_check_prepared.startswith(value_from_rule)
    elif id_type == "contains":
        return value_from_rule in val_to_check_prepared
    elif id_type == "exactmatch":
        return val_to_check_prepared == value_from_rule
    elif id_type == "regex":
        # compiled_regex should be valid at this point
        return bool(compiled_regex.search(value_to_check_str)) # Regex uses original case string
    else:
        logger.warning(f"Unknown identifier type: '{id_type}' in rule: {identifier_rule}")
        return False


# --- Function to Read Processed Excel Data ---
def read_comparison_data(filename: str) -> bool:
    """
    Reads data from '* Comparison' sheets and 'Metadata' sheet
    of a processed Excel file into the Flask app's config cache.
    Uses headers from sheet as keys for row data dictionaries.
    Reads the maximum numeric IDs (DN and AG) from the 'Metadata' sheet.

    Args:
        filename: Path to the processed Excel file (*_processed.xlsx).

    Returns:
        True if data loading was successful (even if no comparison sheets found),
        False if a critical error occurred (e.g., file not found, invalid format).
    """
    # This function modifies current_app.config directly.
    # Ensure it's called within a Flask application context.
    comparison_data_from_excel = {} # Data from comparison sheets
    workbook = None
    comparison_sheet_names_found = [] # List of actual comparison sheet names found
    sheet_headers_cache = {} # Store headers read from each sheet
    max_dn_id_from_metadata = 0 # Initialize max IDs read from file
    max_ag_id_from_metadata = 0

    try:
        logger.info(f"Reading comparison data from processed file: {filename}")
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Processed Excel file not found at {filename}")

        workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        logger.info(f"Workbook '{filename}' loaded successfully. Sheets: {workbook.sheetnames}")

        # --- Read Max IDs from Metadata sheet ---
        if METADATA_SHEET_NAME in workbook.sheetnames:
            try:
                metadata_sheet = workbook[METADATA_SHEET_NAME]
                # Read DN Max ID value
                dn_id_val = metadata_sheet[MAX_DN_ID_VALUE_CELL].value
                if dn_id_val is not None and str(dn_id_val).isdigit():
                    max_dn_id_from_metadata = int(dn_id_val)
                    logger.info(f"Read Max DN ID from '{METADATA_SHEET_NAME}' ({MAX_DN_ID_VALUE_CELL}): {max_dn_id_from_metadata}")
                else:
                    logger.warning(f"Value in '{METADATA_SHEET_NAME}' cell {MAX_DN_ID_VALUE_CELL} is not a valid number: '{dn_id_val}'. Using 0.")

                # Read Agent Group Max ID value
                ag_id_val = metadata_sheet[MAX_AG_ID_VALUE_CELL].value
                if ag_id_val is not None and str(ag_id_val).isdigit():
                    max_ag_id_from_metadata = int(ag_id_val)
                    logger.info(f"Read Max AG ID from '{METADATA_SHEET_NAME}' ({MAX_AG_ID_VALUE_CELL}): {max_ag_id_from_metadata}")
                else:
                     logging.warning(f"Value in '{METADATA_SHEET_NAME}' cell {MAX_AG_ID_VALUE_CELL} is not a valid number: '{ag_id_val}'. Using 0.")
            except Exception as meta_e:
                logger.error(f"Error reading Max IDs from '{METADATA_SHEET_NAME}' sheet: {meta_e}. Using 0 for both.")
        else:
            logger.warning(f"'{METADATA_SHEET_NAME}' sheet not found in workbook '{filename}'. Max IDs will be 0.")

        # Store the read (or default 0) max IDs in app config
        current_app.config['MAX_DN_ID'] = max_dn_id_from_metadata
        current_app.config['MAX_AG_ID'] = max_ag_id_from_metadata
        # --- End Read Max IDs ---


        # Find all sheets ending with the comparison suffix
        comparison_sheet_names_found = sorted([s for s in workbook.sheetnames if s.endswith(COMPARISON_SUFFIX)])
        logger.info(f"Found comparison sheets: {comparison_sheet_names_found}")

        # If no comparison sheets found, still return True but with empty data
        if not comparison_sheet_names_found:
            logging.warning(f"No sheets ending with '{COMPARISON_SUFFIX}' found in {filename}.")
            current_app.config['EXCEL_DATA'] = {}
            current_app.config['COMPARISON_SHEETS'] = []
            current_app.config['EXCEL_FILENAME'] = filename # Store name of loaded file
            current_app.config['SHEET_HEADERS'] = {}
            return True

        # Process each comparison sheet
        for sheet_name in comparison_sheet_names_found:
            sheet = workbook[sheet_name]
            data_rows: List[Dict[str, Any]] = [] # Using a more descriptive name
            try:
                # Read the header row (expected to be row 1)
                headers = [cell.value for cell in sheet[1]]
                # Filter out None headers, ensure they are strings and stripped
                headers = [str(h).strip() for h in headers if h is not None]
                if not headers:
                    raise IndexError("No valid headers found in row 1.")
                sheet_headers_cache[sheet_name] = headers # Cache headers for this sheet
            except IndexError:
                 # Handle case where sheet might be completely empty or has no header
                 logging.warning(f"Sheet '{sheet_name}' seems empty or has no header row. Skipping.")
                 continue # Skip this sheet

            # Read data rows (starting from row 2)
            # Use the length of actual headers read to determine max columns to read
            max_cols = len(headers)
            for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2, max_col=max_cols), start=2):
                row_values = [cell.value for cell in row_cells]
                # Only add row if the first cell (Key/Item) has a value
                if row_values and row_values[0] is not None and str(row_values[0]).strip() != "":
                    # Create dict using the actual headers read as keys
                    row_data_dict = {headers[i]: row_values[i] if i < len(row_values) else None for i in range(max_cols)}
                    # Add the 'Header' key for display purposes in the template (using the first actual header)
                    row_data_dict['Header'] = headers[0]
                    data_rows.append(row_data_dict)

            comparison_data_from_excel[sheet_name] = data_rows # Store data for this sheet
            logging.info(f"Read {len(data_rows)} valid rows from sheet '{sheet_name}'. Headers used as keys: {headers}")

        # --- Store results in app config ---
        current_app.config['EXCEL_DATA'] = comparison_data_from_excel
        current_app.config['COMPARISON_SHEETS'] = comparison_sheet_names_found
        current_app.config['EXCEL_FILENAME'] = filename # Store name of loaded file
        current_app.config['SHEET_HEADERS'] = sheet_headers_cache # Store the read headers
        # MAX_IDs were already stored earlier from Metadata sheet
        # --- End Store results ---

        return True # Indicate success

    except FileNotFoundError:
        logging.error(f"Excel file not found: {filename}")
        # Reset cache on error
        current_app.config['EXCEL_DATA'] = {}
        current_app.config['COMPARISON_SHEETS'] = []
        current_app.config['EXCEL_FILENAME'] = None
        current_app.config['MAX_DN_ID'] = 0
        current_app.config['MAX_AG_ID'] = 0
        current_app.config['SHEET_HEADERS'] = {}
        return False # Indicate failure
    except InvalidFileException:
        logging.error(f"Invalid Excel file format or corrupted file: {filename}")
        current_app.config['EXCEL_DATA'] = {}
        current_app.config['COMPARISON_SHEETS'] = []
        current_app.config['EXCEL_FILENAME'] = None
        current_app.config['MAX_DN_ID'] = 0
        current_app.config['MAX_AG_ID'] = 0
        current_app.config['SHEET_HEADERS'] = {}
        return False # Indicate failure
    except Exception as e:
        # Catch-all for other errors during file processing
        logging.error(f"Error reading Excel file '{filename}': {e}", exc_info=True)
        current_app.config['EXCEL_DATA'] = {}
        current_app.config['COMPARISON_SHEETS'] = []
        current_app.config['EXCEL_FILENAME'] = None
        current_app.config['MAX_DN_ID'] = 0
        current_app.config['MAX_AG_ID'] = 0
        current_app.config['SHEET_HEADERS'] = {}
        return False # Indicate failure
    finally:
        # Ensure workbook is closed to release resources
        if workbook:
            try:
                workbook.close()
                logging.debug("Workbook closed after reading.")
            except Exception as close_e:
                logging.warning(f"Error closing workbook: {close_e}")

