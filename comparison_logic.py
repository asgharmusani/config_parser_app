# -*- coding: utf-8 -*-
"""
Handles the comparison logic between processed Excel data and fetched API data,
and writes the results to dedicated comparison sheets in the workbook.
This module dynamically creates comparison sheets based on the entities
processed by the rule engine.
"""

import logging
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import cell as openpyxl_cell_utils # For get_column_letter
from typing import Dict, Any, Set

logger = logging.getLogger(__name__) # Use module-specific logger

# --- Comparison and Reporting ---
def write_comparison_sheets(
    workbook: openpyxl.workbook.Workbook,
    sheet_data_for_comparison: Dict[str, Set[str]],
    api_data: Dict[str, Dict[str, Any]], # API data structure varies by entity type
    intermediate_data: Dict[str, Dict[str, Dict[str, Any]]] # Full sheet data details
):
    """
    Compares sheet data (non-struck only) with API data and writes results
    to dedicated comparison sheets in the workbook.
    Dynamically handles column layout based on entity type (e.g., Skill Expressions).

    Args:
        workbook: The openpyxl Workbook object to write results into.
        sheet_data_for_comparison: Dict containing sets of non-struck primary keys
                                   for each entity type, as prepared by excel_processing.py.
                                   Format: {"EntityName1": {set_of_keys}, "EntityName2": {set}, ...}
        api_data: Dict containing items fetched from the API, filtered by relevant entity names.
                  Format: {"EntityName1": {key: id_or_details_dict}, ...}
                  For skill_exprs, details_dict is {'id':.., 'expr':.., 'ideal':..}.
                  For others, it's typically {key: id_string}.
        intermediate_data: Dict containing detailed items found in sheets (after rule processing
                           and strike-through resolution). Used to get details for items
                           marked as 'New in Sheet'.
                           Format: {"EntityName1": {key: details_dict_from_sheet}, ...}
    """
    logging.info("Starting comparison and writing results to comparison sheets.")

    # Basic checks for empty data
    if not api_data and not sheet_data_for_comparison:
        logging.warning("Both API data and Sheet data for comparison are empty. No comparison sheets will be generated.")
        return
    if not sheet_data_for_comparison:
        logging.warning("Sheet data for comparison is empty. Comparison sheets might only show 'Missing from Sheet'.")
        # Proceed, as API might have data not in sheet
    if not api_data:
        logging.warning("API data is empty. Comparison sheets might only show 'New in Sheet'.")
        # Proceed, as sheet might have data not in API

    # Determine which entity types to process based on the union of keys from both data sources
    all_entity_keys_to_compare = set(sheet_data_for_comparison.keys()).union(set(api_data.keys()))
    if not all_entity_keys_to_compare:
        logging.info("No common or unique entity keys found in sheet data or API data. Skipping comparison sheet generation.")
        return

    # Iterate through each entity type found
    for entity_name in sorted(list(all_entity_keys_to_compare)): # Process in a consistent order
        # Use the entity_name (from the rule template) as the base for the sheet title
        sheet_title_prefix = entity_name
        logging.info(f"Generating comparison sheet for entity: '{entity_name}'")
        comparison_sheet_title = f"{sheet_title_prefix} Comparison"

        # Ensure sheet doesn't already exist (should have been removed by excel_processing.py)
        if comparison_sheet_title in workbook.sheetnames:
            try:
                del workbook[comparison_sheet_title]
                logging.debug(f"Removed pre-existing sheet: {comparison_sheet_title}")
            except Exception as e:
                 logging.warning(f"Could not remove existing sheet '{comparison_sheet_title}': {e}")

        # Create the new comparison sheet
        sheet = workbook.create_sheet(title=comparison_sheet_title)

        # --- Prepare data specific to this entity type ---
        # Get non-struck item keys from sheet processing for this entity
        sheet_items_non_struck = sheet_data_for_comparison.get(entity_name, set())
        # Get API items for this entity (could be a dict of {key:id} or {key:details_dict})
        api_items_dict = api_data.get(entity_name, {})
        api_items_keys = set(api_items_dict.keys()) # Get all keys (identifiers) from API data

        # Calculate differences based on the primary identifying KEYS
        # Items present (non-struck) in sheet but not present in API
        new_in_sheet = sheet_items_non_struck - api_items_keys
        # Items present in API but not present (non-struck) in sheet
        missing_from_sheet_non_struck = api_items_keys - sheet_items_non_struck

        row_num = 2 # Start writing data rows from row 2

        # --- Set Headers and Column Widths based on entity type ---
        # Heuristic to check if this entity is a "skill expression" type by its name.
        # This relies on the 'name' field in the excelrule_template.json.
        # A more robust method might involve a specific flag in the rule definition.
        is_skill_expression_type = "expression" in entity_name.lower() or \
                                   "skill_expr" in entity_name.lower()

        if is_skill_expression_type:
            # Define headers for the 5-column Skill Exprs comparison sheet
            headers = ["Concatenated Key", "Expression", "Ideal Expression", "ID (from API)", "Status"]
            # Define approximate column widths for better viewing
            col_widths = [45, 45, 35, 20, 35]
        else:
            # Define headers for the standard 3-column comparison sheets
            # Use the entity_name (which was sheet_title_prefix) as the first column header
            headers = [entity_name, "ID (from API)", "Status"]
            col_widths = [45, 20, 35]

        # Write headers to the sheet and apply formatting
        for col_idx, header_text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(bold=True) # Make headers bold
            # Set column width for better readability
            try:
                column_letter = openpyxl_cell_utils.get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].width = col_widths[col_idx-1]
            except IndexError: # Safety check for col_widths definition
                 pass # Ignore error if width definition is wrong


        # --- Write Data Rows ---

        # Write items that are "New in Sheet"
        if new_in_sheet:
            logging.debug(f"'{entity_name}' - Found {len(new_in_sheet)} items New in Sheet (Non-Struck).")
            # Sort items alphabetically by key for consistent report order
            for item_key in sorted(list(new_in_sheet)):
                if is_skill_expression_type:
                    # Lookup details from intermediate_data (which originates from sheet processing)
                    # intermediate_data contains the fully resolved data for items found in the sheet.
                    item_details_from_sheet = intermediate_data.get(entity_name, {}).get(item_key, {})
                    sheet.cell(row=row_num, column=1, value=item_key) # Concatenated Key
                    sheet.cell(row=row_num, column=2, value=item_details_from_sheet.get('expr', item_details_from_sheet.get('Expression',''))) # Expression from sheet
                    sheet.cell(row=row_num, column=3, value=item_details_from_sheet.get('ideal', item_details_from_sheet.get('Ideal Expression',''))) # Ideal Expression from sheet
                    sheet.cell(row=row_num, column=4, value="N/A") # ID (Not applicable as it's not from API)
                    sheet.cell(row=row_num, column=5, value="New in Sheet (Non-Struck)") # Status
                else:
                    # Standard 3-column layout for VQ, Skill, VAG
                    sheet.cell(row=row_num, column=1, value=item_key) # Item Name
                    sheet.cell(row=row_num, column=2, value="N/A") # ID
                    sheet.cell(row=row_num, column=3, value="New in Sheet (Non-Struck)") # Status
                row_num += 1
        else:
             # Log if no items were found only in the sheet
             logging.debug(f"'{entity_name}' - No items found only in the sheet (non-struck).")

        # Write items that are "Missing from Sheet" (or only struck out)
        if missing_from_sheet_non_struck:
             logging.debug(f"'{entity_name}' - Found {len(missing_from_sheet_non_struck)} items Missing from Sheet (or only Struck Out).")
             # Sort items alphabetically by key for consistent report order
             for item_key in sorted(list(missing_from_sheet_non_struck)):
                if is_skill_expression_type:
                    # Lookup details from api_data (which originates from API)
                    # For skill_exprs, api_items_dict[item_key] is a dict: {'id': ..., 'expr': ..., 'ideal': ...}
                    api_item_details = api_items_dict.get(item_key, {})
                    sheet.cell(row=row_num, column=1, value=item_key) # Concatenated Key
                    sheet.cell(row=row_num, column=2, value=api_item_details.get('expr', '')) # Expression from API
                    sheet.cell(row=row_num, column=3, value=api_item_details.get('ideal', '')) # Ideal Expression from API
                    sheet.cell(row=row_num, column=4, value=api_item_details.get('id', 'ID Not Found')) # ID from API
                    sheet.cell(row=row_num, column=5, value="Missing in Sheet (or only Struck Out)") # Status
                else:
                    # Standard 3-column layout for VQ, Skill, VAG
                    # For these, api_items_dict[item_key] is just the ID string
                    api_id_value = api_items_dict.get(item_key, "ID Not Found")
                    sheet.cell(row=row_num, column=1, value=item_key) # Item Name
                    sheet.cell(row=row_num, column=2, value=api_id_value) # ID from API
                    sheet.cell(row=row_num, column=3, value="Missing in Sheet (or only Struck Out)") # Status
                row_num += 1
        else:
            # Log if no items were found only in the API data
            logging.debug(f"'{entity_name}' - No items found only in the API (when compared to non-struck sheet items).")

        logging.info(f"Finished comparison sheet for: {entity_name}")

